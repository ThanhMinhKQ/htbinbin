from fastapi import APIRouter, Request, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload # Thêm joinedload
import io
import openpyxl
from openpyxl.utils import get_column_letter
# --- THÊM CÁC IMPORT NÀY TỪ CALENDAR.PY ---
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
import calendar
from datetime import datetime, date, timedelta
# ------------------------------------------
from urllib.parse import quote

from ..db.session import get_db
from ..db.models import Task, User, AttendanceRecord, Branch, ServiceRecord
from ..core.utils import VN_TZ, format_datetime_display
from ..core.config import logger

# Import các hàm query cũ (giữ nguyên)
from .tasks import _get_filtered_tasks_query
from .results import _get_filtered_records_query

router = APIRouter()

def _auto_adjust_worksheet_columns(worksheet):
    """Helper function to adjust column widths of a worksheet."""
    for i, column_cells in enumerate(worksheet.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        # Also check header length
        header_cell = worksheet.cell(row=1, column=i)
        if header_cell.value:
            max_length = len(str(header_cell.value))

        for cell in column_cells:
            if cell.row == 1: continue # Skip header
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

@router.get("/api/tasks/export-excel", tags=["Export"])
async def export_tasks_to_excel(
    request: Request,
    chi_nhanh: str = "",
    search: str = "",
    trang_thai: str = "",
    han_hoan_thanh: str = "",
    bo_phan: str = "",  # <--- THÊM THAM SỐ NÀY
    db: Session = Depends(get_db)
):
    user_data = request.session.get("user")
    if not user_data or user_data.get("role") not in ["admin", "boss"]:
        raise HTTPException(status_code=403, detail="Bạn không có quyền truy cập.")

    # Truyền bo_phan vào hàm query chung
    tasks_query = _get_filtered_tasks_query(
        db, 
        user_data, 
        chi_nhanh, 
        search, 
        trang_thai, 
        han_hoan_thanh, 
        bo_phan  # <--- TRUYỀN VÀO ĐÂY
    )
    
    # Sắp xếp giống như màn hình danh sách để đồng bộ
    rows_all = tasks_query.order_by(Task.due_date.nullslast()).all()
    
    if not rows_all:
        return Response(status_code=204, content="Không có dữ liệu để xuất.")

    data_for_export = [{
        "ID": t.id_task, # Nên dùng ID Task (mã định danh) thay vì ID database
        "Chi Nhánh": t.branch.branch_code if t.branch else '', # Dùng mã chi nhánh cho gọn
        "Vị trí": t.room_number,
        "Bộ Phận": t.department or "", # <--- THÊM CỘT NÀY VÀO EXCEL
        "Mô Tả": t.description,
        "Ngày Tạo": format_datetime_display(t.created_at, with_time=True),
        "Hạn Hoàn Thành": format_datetime_display(t.due_date, with_time=True), # Có giờ để chính xác hơn
        "Trạng Thái": t.status,
        "Người Tạo": t.author.name if t.author else '',
        "Người Thực Hiện": t.assignee.name if t.assignee else '',
        "Ngày Hoàn Thành": format_datetime_display(t.completed_at, with_time=True) if t.completed_at else "",
        "Ghi Chú": t.notes or "",
    } for t in rows_all]

    # ... (Phần code tạo Excel bên dưới giữ nguyên) ...
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CongViec"

    if data_for_export:
        headers = list(data_for_export[0].keys())
        ws.append(headers)
        for row_data in data_for_export:
            ws.append(list(row_data.values()))
            
    _auto_adjust_worksheet_columns(ws)

    wb.save(output)
    output.seek(0)

    # Lấy kích thước của file trong memory để thêm vào header Content-Length
    file_size = output.getbuffer().nbytes

    filename = f"danh_sach_cong_viec_{datetime.now(VN_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
    # Mã hóa tên file để tương thích với nhiều trình duyệt hơn
    encoded_filename = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Content-Length": str(file_size)
        }
    )

@router.get("/api/attendance/export-excel", tags=["Export"])
async def export_attendance_to_excel(request: Request, db: Session = Depends(get_db)):
    user_data = request.session.get("user")
    if not user_data or user_data.get("role") not in ["admin", "boss"]:
        raise HTTPException(status_code=403, detail="Bạn không có quyền truy cập.")

    # Sử dụng hàm query đã được module hóa từ results.py
    query, columns = _get_filtered_records_query(db, request.query_params, user_data)
    records = db.execute(query).all()

    if not records:
        return Response(status_code=204, content="Không có dữ liệu để xuất.")

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DiemDanh"

    # Ghi header
    ws.append([col.name for col in columns])
    # Ghi dữ liệu
    for rec in records:
        ws.append([format_datetime_display(val) if isinstance(val, datetime) else val for val in rec])
    
    _auto_adjust_worksheet_columns(ws)
    wb.save(output)
    output.seek(0)
    
    # Lấy kích thước của file trong memory để thêm vào header Content-Length
    file_size = output.getbuffer().nbytes
    
    filename = f"ket_qua_diem_danh_{datetime.now(VN_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
    # Mã hóa tên file để tương thích với nhiều trình duyệt hơn
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Content-Length": str(file_size)
        }
    )

# Trong file export.py

@router.get("/api/attendance/calendar-export-excel", tags=["Export"])
def export_attendance_calendar_excel(
    request: Request,
    db: Session = Depends(get_db),
    month: int = None,
    year: int = None,
):
    """
    Xuất Excel: 
    - Sheet 1 (Chấm công): Căn lề trái cho Tên NV và chữ "Tăng ca".
    - Sheet 2 (Dịch vụ): Giữ nguyên (1 dòng gộp).
    """
    user_data = request.session.get("user")
    if not user_data or user_data.get("role") not in ["admin", "boss"]:
        raise HTTPException(status_code=403, detail="Bạn không có quyền truy cập.")

    now = datetime.now(VN_TZ)
    current_month = month if month else now.month
    current_year = year if year else now.year

    start_date_of_month = date(current_year, current_month, 1)
    _, num_days = calendar.monthrange(current_year, current_month)
    end_date_of_month_query = date(current_year, current_month, num_days) + timedelta(days=1)

    # === 1. LẤY DỮ LIỆU (Giữ nguyên) ===
    all_users = db.query(User).options(
        joinedload(User.department), 
        joinedload(User.main_branch)
    ).filter(
        User.employee_code.notin_(['admin', 'boss'])
    ).all()

    def get_sort_key(user):
        branch_code = user.main_branch.branch_code if user.main_branch else 'ZZZ'
        role_code = user.department.role_code if user.department else 'z'
        if branch_code.startswith('B') and branch_code[1:].isdigit():
            branch_sort_key = (0, int(branch_code[1:]))
        else:
            branch_sort_key = (1, branch_code)
        role_priority = {"letan": 0, "buongphong": 1, "baove": 2, "ktv": 3, "quanly": 4}
        role_sort_key = role_priority.get(role_code, 99)
        return (branch_sort_key, role_sort_key, user.name)

    all_users.sort(key=get_sort_key)
    user_main_branch_map = {u.employee_code: u.main_branch.branch_code if u.main_branch else '' for u in all_users}

    housekeeping_users = [u for u in all_users if u.department and 
                          ("buồng" in str(u.department.role_code).lower() or 
                           "buong" in str(u.department.role_code).lower())]

    all_att_records = db.query(AttendanceRecord).options(joinedload(AttendanceRecord.branch)).filter(
        AttendanceRecord.attendance_datetime >= start_date_of_month,
        AttendanceRecord.attendance_datetime < end_date_of_month_query
    ).all()

    all_svc_records = db.query(ServiceRecord).filter(
        ServiceRecord.service_datetime >= start_date_of_month,
        ServiceRecord.service_datetime < end_date_of_month_query
    ).all()

    # === 4. XỬ LÝ DỮ LIỆU (Giữ nguyên logic Pivot cũ) ===
    data_pivot_att = defaultdict(lambda: defaultdict(lambda: {"main_work": 0.0, "overtime_work": 0.0}))
    for rec in all_att_records:
        dt_local = rec.attendance_datetime.astimezone(VN_TZ)
        work_date = dt_local.date() - timedelta(days=1) if dt_local.hour < 7 else dt_local.date()
        if work_date.month != current_month or work_date.year != current_year: continue
        
        day_num = work_date.day
        emp_code = rec.employee_code_snapshot
        main_branch = user_main_branch_map.get(emp_code)
        work_branch = rec.branch.branch_code if rec.branch else ''
        work_units = rec.work_units or 0
        is_ot_branch = (main_branch and work_branch and work_branch != main_branch)

        if rec.is_overtime or is_ot_branch:
            data_pivot_att[emp_code][day_num]["overtime_work"] += work_units
        else:
            data_pivot_att[emp_code][day_num]["main_work"] += work_units

    data_pivot_svc = defaultdict(lambda: defaultdict(lambda: {"Total": 0}))
    for svc in all_svc_records:
        dt_local = svc.service_datetime.astimezone(VN_TZ)
        work_date = dt_local.date() - timedelta(days=1) if dt_local.hour < 7 else dt_local.date()
        if work_date.month != current_month or work_date.year != current_year: continue
        
        day_num = work_date.day
        emp_code = svc.employee_code_snapshot
        qty = svc.quantity or 0
        s_type = (svc.service_type or "").lower()
        if any(k in s_type for k in ['giặt', 'giat', 'ủi', 'ui', 'là']):
            data_pivot_svc[emp_code][day_num]["Total"] += qty

    # === 5. VẼ EXCEL ===
    wb = openpyxl.Workbook()
    
    # --- STYLE ---
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # [THÊM MỚI] Style căn lề trái
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    main_work_fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
    ot_work_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    service_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") 

    cn_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cn_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ["STT", "TÊN NHÂN VIÊN"] + [f"{d:02d}" for d in range(1, num_days + 1)] + ["TỔNG CỘNG"]

    def draw_sheet(ws, sheet_title, data_source, row1_label, row2_label, row1_key, row2_key, row1_fill, row2_fill, user_list, is_single_row=False):
        ws.title = sheet_title
        ws.append(headers)

        for col_idx, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            if col_idx > 2 and col_idx <= num_days + 2:
                day_num = col_idx - 2
                day_date = date(current_year, current_month, day_num)
                if day_date.weekday() == 6: 
                    cell.fill = cn_fill
                    cell.font = cn_font
                    cell.value = f"CN\n{day_num:02d}"

        current_row = 2
        stt = 1
        
        for user in user_list:
            emp_code = user.employee_code
            emp_name_with_branch = f"{user.name}_{user.main_branch.branch_code}" if user.main_branch else user.name
            
            if is_single_row:
                # --- CHẾ ĐỘ 1 DÒNG (SHEET DỊCH VỤ) ---
                display_name = f"{emp_name_with_branch} - {row1_label}" if row1_label else emp_name_with_branch
                row_data = [stt, display_name]
                total_val = 0.0

                for d in range(1, num_days + 1):
                    day_data = data_source[emp_code].get(d, {})
                    val = day_data.get(row1_key, 0)
                    row_data.append(val if val > 0 else "")
                    total_val += val
                
                row_data.append(total_val if total_val > 0 else "")
                ws.append(row_data)

                ws_row = ws[current_row]
                for col_idx in range(1, len(headers) + 1):
                    cell = ws_row[col_idx - 1]
                    cell.border = thin_border
                    cell.alignment = center_align
                    if col_idx > 2 and col_idx <= num_days + 2:
                        cell.fill = row1_fill
                
                # Căn lề trái cho tên
                ws.cell(row=current_row, column=2).alignment = left_align
                current_row += 1
            
            else:
                # --- CHẾ ĐỘ 2 DÒNG (SHEET CHẤM CÔNG) ---
                row1_data = [stt, emp_name_with_branch]
                row2_data = ["", row2_label]
                total_row1 = 0.0
                total_row2 = 0.0

                for d in range(1, num_days + 1):
                    day_data = data_source[emp_code].get(d, {})
                    val1 = day_data.get(row1_key, 0)
                    val2 = day_data.get(row2_key, 0)
                    row1_data.append(val1 if val1 > 0 else "")
                    row2_data.append(val2 if val2 > 0 else "")
                    total_row1 += val1
                    total_row2 += val2
                row1_data.append(total_row1 if total_row1 > 0 else "")
                row2_data.append(total_row2 if total_row2 > 0 else "")

                ws.append(row1_data)
                ws.append(row2_data)

                # Style cơ bản
                ws_row1 = ws[current_row]
                ws_row2 = ws[current_row + 1]

                for col_idx in range(1, len(headers) + 1):
                    cell1 = ws_row1[col_idx - 1]
                    cell2 = ws_row2[col_idx - 1]
                    cell1.border = thin_border
                    cell2.border = thin_border
                    cell1.alignment = center_align # Mặc định căn giữa các ô dữ liệu
                    cell2.alignment = center_align
                    if col_idx > 2 and col_idx <= num_days + 2:
                        cell1.fill = row1_fill
                        cell2.fill = row2_fill

                # Merge STT
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=1)
                ws.cell(row=current_row, column=1).alignment = center_align

                # --- [SỬA] CĂN LỀ TRÁI CHO TÊN VÀ TĂNG CA ---
                # Dòng 1: Tên nhân viên
                name_cell = ws.cell(row=current_row, column=2)
                name_cell.alignment = left_align # <--- Căn trái
                
                # Dòng 2: Chữ "Tăng ca"
                label_cell = ws.cell(row=current_row + 1, column=2)
                label_cell.alignment = left_align # <--- Căn trái
                # ---------------------------------------------

                current_row += 2
            
            stt += 1

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        for d in range(1, num_days + 1):
            ws.column_dimensions[get_column_letter(d + 2)].width = 5
        ws.column_dimensions[get_column_letter(num_days + 3)].width = 10
        ws.freeze_panes = "C2"

    # === 6. VẼ SHEET ===
    ws1 = wb.active
    draw_sheet(
        ws=ws1, 
        sheet_title=f"Chấm công T{current_month}", 
        data_source=data_pivot_att, 
        row1_label="",
        row2_label="Tăng ca", 
        row1_key="main_work", 
        row2_key="overtime_work",
        row1_fill=main_work_fill,
        row2_fill=ot_work_fill,
        user_list=all_users,
        is_single_row=False
    )

    ws2 = wb.create_sheet()
    draw_sheet(
        ws=ws2,
        sheet_title=f"Dịch vụ T{current_month}",
        data_source=data_pivot_svc,
        row1_label="Giặt & Ủi",
        row2_label="",
        row1_key="Total",
        row2_key="",
        row1_fill=service_fill,
        row2_fill=None,
        user_list=housekeeping_users,
        is_single_row=True
    )

    # === 7. XUẤT FILE ===
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    
    filename = f"ChamCong_DichVu_Thang_{current_month}_{current_year}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        output_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
    )
