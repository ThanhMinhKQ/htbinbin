# app/api/pms/pms_export.py
"""
PMS DKLT Export - Xuất file Đăng ký lưu trú (Khai báo tạm trú) cho khách đang ở.

Tách 2 nhóm theo loại giấy tờ:
- VN: id_type ∈ {cccd, cmnd, gplx, NULL} — xuất theo mẫu DS_KHACH_VIET_NAM_LUU_TRU (Excel).
- Nước ngoài: id_type ∈ {passport, visa} — xuất theo mẫu KHAI_BAO_TAM_TRU (XML hoặc Excel).

Phòng giờ (stay.check_out_at IS NULL) được loại khỏi cả hai nhóm.
"""
from __future__ import annotations

import io
import json
import re
import unicodedata
import xml.etree.ElementTree as ET
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Optional, List, Dict, Any
from urllib.parse import quote

from fastapi import APIRouter, Body, Depends, HTTPException, Request
from fastapi.responses import JSONResponse, StreamingResponse
from sqlalchemy.orm import Session, joinedload, selectinload

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ...core.utils import VN_TZ
from ...db.models import Branch, HotelGuest, HotelRoom, HotelStay, HotelStayStatus
from ...db.session import get_db
from .pms_helpers import _require_login, _is_admin, _active_branch
from .vn_address import convert_old_to_new_sync

router = APIRouter()


FOREIGN_ID_TYPES = {"passport", "visa"}

# ─── Mẫu nước ngoài (KHAI_BAO_TAM_TRU) ───
FOREIGN_FIELDS = [
    "so_thu_tu",
    "ho_ten",
    "ngay_sinh",
    "ngay_sinh_dung_den",
    "gioi_tinh",
    "ma_quoc_tich",
    "so_ho_chieu",
    "so_phong",
    "ngay_den",
    "ngay_di_du_kien",
    "ngay_tra_phong",
]

# ─── Mẫu Việt Nam (DS_KHACH_VIET_NAM_LUU_TRU) ───
VN_HEADERS = [
    "STT",
    "HỌ TÊN (*)",
    "NGÀY SINH (*)",
    "GIỚI TÍNH (*)",
    "QUỐC TỊCH (*)",
    "LOẠI GIẤY TỜ (*)",
    "TÊN GIẤY TỜ",
    "SỐ GIẤY TỜ (*)",
    "SỐ ĐIỆN THOẠI",
    "NƠI CƯ TRÚ HIỆN NAY",
    "TỈNH/ THÀNH PHỐ",
    "PHƯỜNG/ XÃ/ ĐẶC KHU",
    "ĐỊA CHỈ CHI TIẾT",
    "NGÀY ĐẾN (*)",
    "NGÀY ĐI DỰ KIẾN (*)",
    "SỐ PHÒNG/ KHOA",
    "LÝ DO CƯ TRÚ (*)",
    "NHẬP LÝ DO",
    "GHI CHÚ",
]

VN_TITLE = (
    "DANH SÁCH THÔNG BÁO LƯU TRÚ CHO CÔNG DÂN VIỆT NAM "
    "(*Vui lòng không xóa dữ liệu mẫu; hãy thêm thông tin từ dòng tiếp theo)"
)

# Row 3 trong file gốc — placeholder "Mẫu dữ liệu" để cổng nhận diện cấu trúc.
VN_TEMPLATE_HINT_ROW = [
    "Mẫu dữ liệu",
    "<Họ và tên gồm chữ cái và ký tự>",
    "<Giá trị: dd/mm/yyyy>",
    "<Giá trị: F - Nữ, M - Nam>",
    "<Chọn Quốc tịch theo  Sheet [DANH_MUC]>",
    "<Chọn Loại Giấy Tờ theo  Sheet [DANH_MUC]>",
    "<Bắt buộc nhập với Giấy tờ khác >",
    "<Bắt buộc nhập: Số giấy tờ>",
    "<Không bắt buộc>",
    "<Chọn Nơi cư trú  Sheet [DANH_MUC]>",
    "<Chọn Tỉnh/Thành phố  Sheet [TINH_THANH]>",
    "<Chọn Phường/Xã/Đặc khu  Sheet [PHUONG_XA]>",
    "<Nhập thông tin địa chỉ chi tiết>",
    "<dd/MM/yyyy>",
    "<dd/MM/yyyy>",
    "< Nhập tên Phòng/Khoa>",
    "<Chọn Lý do cư trú theo  Sheet [DANH_MUC]>",
    "<Bắt buộc nhập với Lý do cư trú là Mục đích khác >",
    "",
]

# Row 4 trong file gốc — dòng ví dụ [EXAMPLE]; cổng skip dòng này khi import.
VN_EXAMPLE_ROW = [
    "[EXAMPLE]",
    "[TEST] NGUYỄN QUANG MINH",
    "10/02/1999",
    "M - Nam",
    "VNM - Viet Nam",
    "1 - Thẻ CCCD",
    "",
    "034099009984",
    "03669541874",
    "1 - Thường trú",
    "101 - TP. Hà Nội",
    "101900565 - Xã Gia Lâm",
    "Ngõ 3/6A",
    "20/03/2026",
    "30/04/2026",
    "205A0",
    "3 - Học tập",
    "",
    "",
]

# Mẫu Foreign Excel (sheet KBTT) — header + dòng ví dụ.
FOREIGN_XLSX_TITLE = "DANH SÁCH HỒ SƠ KBTT"
FOREIGN_XLSX_HEADERS = [
    "STT", "HỌ TÊN ", "NGÀY SINH", "NGÀY SINH ĐÚNG ĐẾN", "GIỚI TÍNH",
    "MÃ QUỐC TỊCH", "SỐ HỘ CHIẾU", "SỐ PHÒNG", "NGÀY ĐẾN ",
    "NGÀY ĐI DỰ KIẾN", "NGÀY TRẢ PHÒNG",
]
FOREIGN_XLSX_EXAMPLE_ROW = [
    1, "[TEST] SAMPLE", "01/01/1990", "D - Ngày", "F - Nữ",
    "VNM - Viet Nam", "[TEST] AVAD123", "P134", "20/10/2026",
    "22/10/2026", "22/10/2026",
]

# id_type (DB) → label theo DANH_MUC
ID_TYPE_LABEL = {
    "cccd":     "1 - Thẻ CCCD",
    "cmnd":     "2 - Thẻ CMND",
    "gplx":     "3 - Giấy phép lái xe",
    "passport": "4 - Hộ chiếu",
    "visa":     "4 - Hộ chiếu",
}

# ─── Lookup tỉnh/phường để xuất đúng format "Mã - Tên" ───
_DATA_DIR = Path(__file__).resolve().parent.parent.parent / "static" / "data"
_PROV_FILE = _DATA_DIR / "dklt_provinces.json"
_WARD_FILE = _DATA_DIR / "dklt_wards.json"


def _strip_accents(s: str) -> str:
    s = unicodedata.normalize("NFKD", s or "")
    return "".join(c for c in s if not unicodedata.combining(c))


def _norm(s: str) -> str:
    s = _strip_accents(s).lower().strip()
    # Fold Unicode punctuation variants so keys match across datasets:
    # conversion map (vn_ward_map.json) and export tables (dklt_wards.json)
    # spell apostrophes/dashes differently (curly vs ASCII, en-dash vs hyphen),
    # and a few conversion entries carry stray backslashes (D\\'Ran). Without
    # folding, wards like "M'Drắk", "Ea H'Leo", "Chân Mây – Lăng Cô" never match.
    s = s.replace("\\", "")
    s = re.sub(r"[‘’ʼ`´]", "'", s)   # apostrophe variants → '
    s = re.sub(r"[‐-―]", "-", s)                     # dash variants → -
    return re.sub(r"\s+", " ", s).strip()


def _norm_prov(s: str) -> str:
    s = _norm(s)
    return re.sub(r"^(tp\.?|thanh pho|tinh)\s+", "", s)


def _norm_ward(s: str) -> str:
    s = _norm(s)
    return re.sub(r"^(phuong|xa|dac khu|thi tran|p\.|x\.)\s+", "", s)


@lru_cache(maxsize=1)
def _load_prov_lookup() -> Dict[str, Any]:
    try:
        with open(_PROV_FILE, encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {"provinces": {}, "prov_code_by_name": {}}


@lru_cache(maxsize=1)
def _load_ward_lookup() -> Dict[str, Dict[str, str]]:
    try:
        with open(_WARD_FILE, encoding="utf-8") as f:
            raw = json.load(f)
    except FileNotFoundError:
        return {}

    # Re-normalize stored keys through the CURRENT _norm so the lookup side and
    # the query side always use identical normalization. The JSON keys were
    # pre-baked by a build script with older rules (e.g. en-dash, curly quotes
    # left unfolded), so without this a runtime _norm fix alone can't reach them.
    # Non-destructive: original keys are kept; folded aliases are added.
    folded: Dict[str, Dict[str, str]] = {}
    for prov_code, bucket in raw.items():
        if not isinstance(bucket, dict):
            folded[prov_code] = bucket
            continue
        new_bucket = dict(bucket)
        for key, display in bucket.items():
            alias = _norm(key)
            new_bucket.setdefault(alias, display)
        folded[prov_code] = new_bucket
    return folded


def _resolve_province_display(raw: Optional[str]) -> tuple[str, Optional[str]]:
    """Map giá trị tỉnh từ DB → ('701 - TP. Hồ Chí Minh', '701').

    Nếu raw đã là 'Mã - Tên' thì giữ nguyên; nếu chỉ có tên thì lookup. Trả về
    (display, code) — code dùng để narrow danh sách phường.
    """
    if not raw:
        return "", None
    raw = raw.strip()
    m = re.match(r"^(\d{3,})\s*-\s*(.+)$", raw)
    if m:
        return raw, m.group(1)

    data = _load_prov_lookup()
    code = data.get("prov_code_by_name", {}).get(_norm_prov(raw))
    if code:
        return data.get("provinces", {}).get(_norm_prov(raw), ""), code

    # Fallback: thử match theo norm thường
    display = data.get("provinces", {}).get(_norm(raw)) or \
              data.get("provinces", {}).get(_norm_prov(raw))
    if display:
        m2 = re.match(r"^(\d+)\s*-", display)
        return display, (m2.group(1) if m2 else None)
    return raw, None


def _resolve_ward_display(raw: Optional[str], prov_code: Optional[str]) -> str:
    """Map giá trị phường/xã từ DB → '101900565 - Xã Gia Lâm'.

    Cần `prov_code` để narrow trong tỉnh (tránh trùng tên giữa các tỉnh).
    """
    if not raw:
        return ""
    raw = raw.strip()
    if re.match(r"^\d{6,}\s*-\s*.+$", raw):
        return raw

    if not prov_code:
        return raw

    bucket = _load_ward_lookup().get(prov_code) or {}
    display = bucket.get(_norm_ward(raw)) or bucket.get(_norm(raw))
    return display or raw


def _has_code(display: str) -> bool:
    """True nếu giá trị đã ở dạng 'Mã - Tên' (cổng yêu cầu mã ở đầu)."""
    return bool(re.match(r"^\d{3,}\s*-\s*.+$", (display or "").strip()))


def _resolve_vn_location(g: HotelGuest) -> tuple[str, str]:
    """Resolve (tỉnh, phường) sang format 'Mã - Tên' mà cổng Bộ Công An yêu cầu.

    Khách check-in trước cải cách 1/7/2025 (hoặc qua OCR/import/nhập tay bỏ qua
    bước convert) có thể lưu phường cũ đã bị sáp nhập — không có trong bảng mã
    mới nên lookup trực tiếp thất bại. Khi đó dùng convert_old_to_new_sync (đã
    dùng ở check-in) để ánh xạ phường cũ → mới rồi resolve lại sang mã.

    Ưu tiên dữ liệu đã chuẩn hóa sẵn trên record; chỉ convert khi cần.
    """
    prov_display, prov_code = _resolve_province_display(g.city)
    ward_display = _resolve_ward_display(g.ward, prov_code)

    if _has_code(ward_display):
        return prov_display, ward_display

    # Direct lookup không ra mã → thử convert phường cũ → mới.
    old_ward = (getattr(g, "old_ward", None) or g.ward or "").strip()
    old_city = (getattr(g, "old_city", None) or g.city or "").strip()
    old_district = (getattr(g, "old_district", None) or getattr(g, "district", None) or "").strip()
    if not old_ward:
        return prov_display, ward_display

    conv = convert_old_to_new_sync(old_ward, old_city, old_district) or {}
    if not conv.get("matched"):
        return prov_display, ward_display

    new_prov_display, new_prov_code = _resolve_province_display(
        conv.get("new_province") or g.city
    )
    new_ward_display = _resolve_ward_display(conv.get("new_ward"), new_prov_code)

    # Chỉ thay khi convert thực sự cho ra mã hợp lệ, tránh làm xấu dữ liệu đúng sẵn.
    final_prov = new_prov_display if _has_code(new_prov_display) else prov_display
    final_ward = new_ward_display if _has_code(new_ward_display) else ward_display
    return final_prov, final_ward


# ─────────────────────────── Helpers ─────────────────────────────

def _resolve_branch(request: Request, branch_id: Optional[int], db: Session) -> Branch:
    """Resolve target branch theo pattern của pms_rooms.py."""
    user = _require_login(request)
    is_admin = _is_admin(user)
    branch_code = _active_branch(request)

    target_branch_id: Optional[int] = None
    if is_admin:
        target_branch_id = branch_id
        if not target_branch_id and branch_code:
            b = db.query(Branch).filter(Branch.branch_code == branch_code).first()
            target_branch_id = b.id if b else None
    else:
        b = db.query(Branch).filter(Branch.branch_code == branch_code).first()
        target_branch_id = b.id if b else None

    if not target_branch_id:
        raise HTTPException(status_code=400, detail="Chưa chọn chi nhánh để xuất ĐKLT.")

    branch = db.query(Branch).filter(Branch.id == target_branch_id).first()
    if not branch:
        raise HTTPException(status_code=404, detail="Không tìm thấy chi nhánh.")
    return branch


def _is_foreign(guest: HotelGuest) -> bool:
    return (guest.id_type or "").strip().lower() in FOREIGN_ID_TYPES


def _format_dmy(value) -> str:
    if not value:
        return ""
    try:
        return value.strftime("%d/%m/%Y")
    except Exception:
        return ""


def _gender_code(g: Optional[str]) -> str:
    """Mã 1 ký tự cho mẫu nước ngoài: M / F / U."""
    if not g:
        return "U"
    s = g.strip().lower()
    if s in ("nam", "male", "m"):
        return "M"
    if s in ("nữ", "nu", "female", "f"):
        return "F"
    return "U"


def _gender_label_vn(g: Optional[str]) -> str:
    """Label đầy đủ cho mẫu VN: 'M - Nam' / 'F - Nữ'."""
    if not g:
        return ""
    s = g.strip().lower()
    if s in ("nam", "male", "m"):
        return "M - Nam"
    if s in ("nữ", "nu", "female", "f"):
        return "F - Nữ"
    return ""


def _nationality_code(nat: Optional[str], is_foreign: bool) -> str:
    """Trích mã 3 ký tự đầu (VD: 'VNM - Việt Nam' → 'VNM')."""
    if not nat:
        return "" if is_foreign else "VNM"
    raw = nat.strip()
    head = raw.split("-")[0].strip().split()[0].strip().upper() if raw else ""
    if len(head) >= 2:
        return head
    return "" if is_foreign else "VNM"


def _nationality_label_vn(nat: Optional[str]) -> str:
    """Label đầy đủ cho mẫu VN. Mặc định khách VN → 'VNM - Viet Nam'."""
    raw = (nat or "").strip()
    if raw:
        return raw.replace("Việt Nam", "Viet Nam")
    return "VNM - Viet Nam"


def _foreign_nationality_label(code: str) -> str:
    """Foreign portal Excel cần dạng đầy đủ 'XXX - Country'.

    Nếu DB đã lưu sẵn 'USA - United States' thì trả nguyên; nếu chỉ có
    'USA' thì trả luôn 'USA' (portal vẫn parse được phần code đầu).
    """
    raw = (code or "").strip()
    if "-" in raw:
        return raw
    return raw


def _id_type_label_vn(id_type: Optional[str]) -> str:
    key = (id_type or "").strip().lower()
    return ID_TYPE_LABEL.get(key, "1 - Thẻ CCCD")


def _query_active_guests(db: Session, branch_id: int) -> List[Dict[str, Any]]:
    """Lấy tất cả khách đang ACTIVE thuộc branch, kèm room và stay info.

    Bỏ qua phòng giờ (không có ngày đi dự kiến — stay.check_out_at IS NULL).
    """
    stays = (
        db.query(HotelStay)
        .options(
            selectinload(HotelStay.guests),
            joinedload(HotelStay.room),
        )
        .filter(
            HotelStay.branch_id == branch_id,
            HotelStay.status == HotelStayStatus.ACTIVE,
            HotelStay.check_out_at.isnot(None),
        )
        .all()
    )
    rows: List[Dict[str, Any]] = []
    for stay in stays:
        room_number = stay.room.room_number if stay.room else ""
        for g in (stay.guests or []):
            if g.check_out_at:
                continue
            rows.append({
                "guest": g,
                "stay": stay,
                "room_number": room_number,
            })
    return rows


# ─────────────────────────── Foreign builders ─────────────────────────────

def _foreign_row(row: Dict[str, Any], seq: int) -> Dict[str, str]:
    g: HotelGuest = row["guest"]
    s: HotelStay = row["stay"]
    return {
        "so_thu_tu": str(seq),
        "ho_ten": (g.full_name or "").strip(),
        "ngay_sinh": _format_dmy(g.birth_date),
        "ngay_sinh_dung_den": "D",
        "gioi_tinh": _gender_code(g.gender),
        "ma_quoc_tich": _nationality_code(g.nationality, True),
        "ma_quoc_tich_full": (g.nationality or "").strip(),
        "so_ho_chieu": (g.cccd or "").strip(),
        "so_phong": row["room_number"] or "",
        "ngay_den": _format_dmy(g.check_in_at or s.check_in_at),
        "ngay_di_du_kien": _format_dmy(s.check_out_at),
        "ngay_tra_phong": _format_dmy(s.check_out_at),
    }


def _build_foreign_xml(rows: List[Dict[str, str]]) -> bytes:
    root = ET.Element("KHAI_BAO_TAM_TRU")
    for r in rows:
        item = ET.SubElement(root, "THONG_TIN_KHACH")
        for tag in FOREIGN_FIELDS:
            ET.SubElement(item, tag).text = r.get(tag, "")
    ET.indent(root, space="    ")
    buf = io.BytesIO()
    buf.write(b'<?xml version="1.0" encoding="UTF-8"?>\n')
    buf.write(ET.tostring(root, encoding="utf-8"))
    return buf.getvalue()


def _build_foreign_xlsx(rows: List[Dict[str, str]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KBTT"

    n_cols = len(FOREIGN_XLSX_HEADERS)
    last_col = get_column_letter(n_cols)

    # Row 1: title merged "DANH SÁCH HỒ SƠ KBTT"
    ws.cell(row=1, column=1, value=FOREIGN_XLSX_TITLE)
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill("solid", fgColor="FFF2CC")
    ws.row_dimensions[1].height = 28

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    # Row 2: header tiếng Việt
    for col_idx, label in enumerate(FOREIGN_XLSX_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = header_font
        cell.alignment = header_align
        cell.fill = header_fill
        cell.border = border
    ws.row_dimensions[2].height = 30

    # Row 3: dòng [TEST] SAMPLE — bắt buộc giữ.
    example_align = Alignment(vertical="center", wrap_text=True)
    example_fill = PatternFill("solid", fgColor="E2EFDA")
    for col_idx, val in enumerate(FOREIGN_XLSX_EXAMPLE_ROW, start=1):
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.alignment = example_align
        cell.fill = example_fill
        cell.border = border

    # Row 4+: data thật. Map từ FOREIGN_FIELDS sang format VN của portal.
    data_align = Alignment(vertical="center", wrap_text=True)
    for r_idx, r in enumerate(rows, start=1):
        excel_row = r_idx + 3
        ws.cell(row=excel_row, column=1, value=int(r.get("so_thu_tu") or r_idx))
        ws.cell(row=excel_row, column=2, value=r.get("ho_ten", ""))
        ws.cell(row=excel_row, column=3, value=r.get("ngay_sinh", ""))
        ws.cell(row=excel_row, column=4, value="D - Ngày")
        gender_label = {"M": "M - Nam", "F": "F - Nữ"}.get(r.get("gioi_tinh", ""), "")
        ws.cell(row=excel_row, column=5, value=gender_label)
        ws.cell(row=excel_row, column=6, value=_foreign_nationality_label(
            r.get("ma_quoc_tich_full") or r.get("ma_quoc_tich", "")
        ))
        ws.cell(row=excel_row, column=7, value=r.get("so_ho_chieu", ""))
        ws.cell(row=excel_row, column=8, value=r.get("so_phong", ""))
        ws.cell(row=excel_row, column=9, value=r.get("ngay_den", ""))
        ws.cell(row=excel_row, column=10, value=r.get("ngay_di_du_kien", ""))
        ws.cell(row=excel_row, column=11, value=r.get("ngay_tra_phong", ""))
        for col_idx in range(1, n_cols + 1):
            ws.cell(row=excel_row, column=col_idx).alignment = data_align
            ws.cell(row=excel_row, column=col_idx).border = border

    widths = [6, 28, 13, 16, 12, 24, 18, 12, 13, 16, 16]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────── VN builder ─────────────────────────────

def _vn_row_values(row: Dict[str, Any], seq: int) -> List[str]:
    """Map 1 dòng theo thứ tự VN_HEADERS."""
    g: HotelGuest = row["guest"]
    s: HotelStay = row["stay"]
    prov_display, ward_display = _resolve_vn_location(g)
    return [
        str(seq),                                                # STT
        (g.full_name or "").strip(),                             # HỌ TÊN
        _format_dmy(g.birth_date),                               # NGÀY SINH
        _gender_label_vn(g.gender),                              # GIỚI TÍNH
        _nationality_label_vn(g.nationality),                    # QUỐC TỊCH
        _id_type_label_vn(g.id_type),                            # LOẠI GIẤY TỜ
        "",                                                       # TÊN GIẤY TỜ
        (g.cccd or "").strip(),                                  # SỐ GIẤY TỜ
        (g.phone or "").strip(),                                 # SỐ ĐIỆN THOẠI
        "1 - Thường trú",                                         # NƠI CƯ TRÚ HIỆN NAY
        prov_display,                                             # TỈNH/THÀNH PHỐ (Mã - Tên)
        ward_display,                                             # PHƯỜNG/XÃ (Mã - Tên)
        (g.address or "").strip(),                               # ĐỊA CHỈ CHI TIẾT
        _format_dmy(g.check_in_at or s.check_in_at),             # NGÀY ĐẾN
        _format_dmy(s.check_out_at),                             # NGÀY ĐI DỰ KIẾN
        row["room_number"] or "",                                # SỐ PHÒNG
        "1 - Du lịch",                                            # LÝ DO CƯ TRÚ
        "",                                                       # NHẬP LÝ DO
        "",                                                       # GHI CHÚ
    ]


def _build_vn_xlsx(rows_raw: List[Dict[str, Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DS_KHACH_VIET_NAM_LUU_TRU"

    n_cols = len(VN_HEADERS)
    last_col = get_column_letter(n_cols)

    # Row 1: title (merged)
    ws.cell(row=1, column=1, value=VN_TITLE)
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_cell.fill = PatternFill("solid", fgColor="FFF2CC")
    ws.row_dimensions[1].height = 36

    # Row 2: header
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for col_idx, label in enumerate(VN_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = header_font
        cell.alignment = header_align
        cell.fill = header_fill
        cell.border = border
    ws.row_dimensions[2].height = 32

    # Row 3: dòng "Mẫu dữ liệu" — placeholder bắt buộc để cổng nhận diện schema.
    hint_align = Alignment(vertical="center", wrap_text=True)
    hint_fill = PatternFill("solid", fgColor="FFF2CC")
    for col_idx, val in enumerate(VN_TEMPLATE_HINT_ROW, start=1):
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.alignment = hint_align
        cell.fill = hint_fill
        cell.border = border
        cell.font = Font(italic=True, color="808080")

    # Row 4: dòng [EXAMPLE] — cổng skip khi import.
    example_align = Alignment(vertical="center", wrap_text=True)
    example_fill = PatternFill("solid", fgColor="E2EFDA")
    for col_idx, val in enumerate(VN_EXAMPLE_ROW, start=1):
        cell = ws.cell(row=4, column=col_idx, value=val)
        cell.alignment = example_align
        cell.fill = example_fill
        cell.border = border

    # Row 5+: data thật
    data_align = Alignment(vertical="center", wrap_text=True)
    for r_idx, row in enumerate(rows_raw, start=1):
        values = _vn_row_values(row, r_idx)
        excel_row = r_idx + 4
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.alignment = data_align
            cell.border = border

    widths = [
        6,   # STT
        26,  # HỌ TÊN
        13,  # NGÀY SINH
        12,  # GIỚI TÍNH
        18,  # QUỐC TỊCH
        20,  # LOẠI GIẤY TỜ
        18,  # TÊN GIẤY TỜ
        18,  # SỐ GIẤY TỜ
        15,  # SỐ ĐIỆN THOẠI
        18,  # NƠI CƯ TRÚ HIỆN NAY
        20,  # TỈNH/THÀNH PHỐ
        24,  # PHƯỜNG/XÃ
        28,  # ĐỊA CHỈ CHI TIẾT
        13,  # NGÀY ĐẾN
        15,  # NGÀY ĐI DỰ KIẾN
        14,  # SỐ PHÒNG
        18,  # LÝ DO CƯ TRÚ
        18,  # NHẬP LÝ DO
        18,  # GHI CHÚ
    ]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _stream_file(content: bytes, filename: str, media_type: str) -> StreamingResponse:
    encoded = quote(filename)
    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded}",
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


# ─────────────────────────── Endpoints ─────────────────────────────

@router.get("/api/pms/dklt/preview", tags=["PMS DKLT Export"])
def api_dklt_preview(
    request: Request,
    branch_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """Đếm số khách VN và nước ngoài đang ở để frontend biết cần tải mấy file."""
    branch = _resolve_branch(request, branch_id, db)
    rows = _query_active_guests(db, branch.id)

    vn_count = sum(1 for r in rows if not _is_foreign(r["guest"]))
    foreign_count = sum(1 for r in rows if _is_foreign(r["guest"]))

    return JSONResponse({
        "vn_count": vn_count,
        "foreign_count": foreign_count,
        "branch_id": branch.id,
        "branch_code": branch.branch_code,
        "branch_name": branch.name,
    })


@router.get("/api/pms/dklt/rooms", tags=["PMS DKLT Export"])
def api_dklt_rooms(
    request: Request,
    branch_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """Liệt kê các stay đang lưu trú của branch để user chọn phòng cần xuất.

    Bỏ phòng giờ (stay.check_out_at IS NULL) như flow export.
    """
    branch = _resolve_branch(request, branch_id, db)
    rows = _query_active_guests(db, branch.id)

    by_stay: Dict[int, Dict[str, Any]] = {}
    for r in rows:
        stay = r["stay"]
        sid = stay.id
        if sid not in by_stay:
            by_stay[sid] = {
                "stay_id": sid,
                "room_number": r["room_number"] or "",
                "check_in_at": stay.check_in_at.isoformat() if stay.check_in_at else None,
                "check_out_at": stay.check_out_at.isoformat() if stay.check_out_at else None,
                "vn_count": 0,
                "foreign_count": 0,
                "primary_guest": "",
                "guests": [],
            }
        bucket = by_stay[sid]
        g = r["guest"]
        if _is_foreign(g):
            bucket["foreign_count"] += 1
        else:
            bucket["vn_count"] += 1
        if not bucket["primary_guest"] and getattr(g, "is_primary", False):
            bucket["primary_guest"] = (g.full_name or "").strip()
        guest_check_in = g.check_in_at or stay.check_in_at
        bucket["guests"].append({
            "guest_id": g.id,
            "full_name": (g.full_name or "").strip(),
            "is_foreign": _is_foreign(g),
            "is_primary": bool(getattr(g, "is_primary", False)),
            "check_in_at": guest_check_in.isoformat() if guest_check_in else None,
            "dklt_exported_at": g.dklt_exported_at.isoformat() if g.dklt_exported_at else None,
        })

    items = list(by_stay.values())
    for it in items:
        if not it["primary_guest"]:
            for r in rows:
                if r["stay"].id == it["stay_id"]:
                    it["primary_guest"] = (r["guest"].full_name or "").strip()
                    break
        it["guests"].sort(key=lambda x: (not x["is_primary"], (x["full_name"] or "").lower()))
    items.sort(key=lambda x: (x["room_number"] or "", x["stay_id"]))

    return JSONResponse({
        "branch_id": branch.id,
        "branch_code": branch.branch_code,
        "branch_name": branch.name,
        "stays": items,
    })


@router.get("/api/pms/dklt/export", tags=["PMS DKLT Export"])
def api_dklt_export(
    request: Request,
    branch_id: Optional[int] = None,
    group: str = "vn",
    format: str = "xml",
    stay_ids: Optional[str] = None,
    guest_ids: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """
    Xuất file ĐKLT.
    - group: 'vn' | 'foreign'
    - format: 'xml' | 'excel'
        + VN chỉ hỗ trợ excel (XML không có mẫu cho công dân VN).
        + Foreign hỗ trợ cả xml và excel.
    - guest_ids: chuỗi id khách phân cách dấu phẩy. Ưu tiên hơn stay_ids nếu có.
    - stay_ids: chuỗi id stay phân cách dấu phẩy. Bỏ trống = tất cả phòng đang ở.
    """
    group_norm = (group or "").strip().lower()
    fmt_norm = (format or "").strip().lower()
    if group_norm not in ("vn", "foreign"):
        raise HTTPException(status_code=400, detail="Tham số group phải là 'vn' hoặc 'foreign'.")
    if fmt_norm not in ("xml", "excel"):
        raise HTTPException(status_code=400, detail="Tham số format phải là 'xml' hoặc 'excel'.")
    if group_norm == "vn" and fmt_norm == "xml":
        raise HTTPException(
            status_code=400,
            detail="Mẫu khai báo công dân VN chỉ hỗ trợ Excel, không có XML.",
        )

    branch = _resolve_branch(request, branch_id, db)
    rows_raw = _query_active_guests(db, branch.id)

    if guest_ids:
        try:
            selected_guest_ids = {int(x) for x in guest_ids.split(",") if x.strip()}
        except ValueError:
            raise HTTPException(status_code=400, detail="guest_ids không hợp lệ.")
        if selected_guest_ids:
            rows_raw = [r for r in rows_raw if r["guest"].id in selected_guest_ids]
    elif stay_ids:
        try:
            selected_ids = {int(x) for x in stay_ids.split(",") if x.strip()}
        except ValueError:
            raise HTTPException(status_code=400, detail="stay_ids không hợp lệ.")
        if selected_ids:
            rows_raw = [r for r in rows_raw if r["stay"].id in selected_ids]

    want_foreign = (group_norm == "foreign")
    filtered = [r for r in rows_raw if _is_foreign(r["guest"]) == want_foreign]

    if not filtered:
        raise HTTPException(status_code=404, detail="Không có khách thuộc nhóm này.")

    filtered.sort(key=lambda r: (
        (r["room_number"] or ""),
        (r["guest"].full_name or "").lower(),
    ))

    ts = datetime.now(VN_TZ).strftime("%Y%m%d_%H%M%S")
    branch_slug = (branch.branch_code or f"branch{branch.id}").strip()
    group_label = "VN" if group_norm == "vn" else "NN"

    if group_norm == "foreign":
        rows = [_foreign_row(r, idx) for idx, r in enumerate(filtered, start=1)]
        if fmt_norm == "xml":
            content = _build_foreign_xml(rows)
            filename = f"DKLT_{group_label}_{branch_slug}_{ts}.xml"
            return _stream_file(content, filename, "application/xml; charset=utf-8")
        content = _build_foreign_xlsx(rows)
        filename = f"DKLT_{group_label}_{branch_slug}_{ts}.xlsx"
        return _stream_file(
            content,
            filename,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # group == "vn" → chỉ excel theo mẫu DS_KHACH_VIET_NAM_LUU_TRU
    content = _build_vn_xlsx(filtered)
    filename = f"DKLT_{group_label}_{branch_slug}_{ts}.xlsx"
    return _stream_file(
        content,
        filename,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.post("/api/pms/dklt/mark-exported", tags=["PMS DKLT Export"])
def api_dklt_mark_exported(
    request: Request,
    payload: Dict[str, Any] = Body(...),
    branch_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """Đánh dấu các khách đã xuất ĐKLT (set dklt_exported_at/by).

    Gọi sau khi frontend tải file thành công. Chỉ cập nhật khách thuộc branch
    đã resolve để tránh ghi chéo chi nhánh.
    """
    user = _require_login(request)
    branch = _resolve_branch(request, branch_id, db)

    raw_ids = payload.get("guest_ids") if isinstance(payload, dict) else None
    if not isinstance(raw_ids, list) or not raw_ids:
        raise HTTPException(status_code=400, detail="guest_ids không hợp lệ.")
    try:
        guest_id_set = {int(x) for x in raw_ids}
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="guest_ids không hợp lệ.")

    user_id = user.get("id")
    now = datetime.now(VN_TZ)

    rows = (
        db.query(HotelGuest)
        .join(HotelStay, HotelGuest.stay_id == HotelStay.id)
        .filter(
            HotelGuest.id.in_(guest_id_set),
            HotelStay.branch_id == branch.id,
        )
        .all()
    )
    for g in rows:
        g.dklt_exported_at = now
        g.dklt_exported_by = user_id
    db.commit()

    return JSONResponse({"updated": len(rows)})

