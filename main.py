import secrets
from fastapi import FastAPI, Request, Form, HTTPException, Depends, BackgroundTasks
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, FileResponse, Response, StreamingResponse
from pydantic import BaseModel
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from utils import parse_datetime_input, format_datetime_display, is_overdue
import atexit
from apscheduler.schedulers.background import BackgroundScheduler
from fastapi import Request
from typing import Optional, List
from config import logger

from database import SessionLocal, get_db, Base
from models import User, Task, AttendanceLog, AttendanceRecord, ServiceRecord, LostAndFoundItem, LostItemStatus
from config import DATABASE_URL, SMTP_CONFIG, ALERT_EMAIL
from database import init_db
from sqlalchemy.orm import Session
from sqlalchemy import or_, cast, Date, desc
from sqlalchemy.exc import SQLAlchemyError
from services.missing_attendance_service import update_missing_attendance_to_db
from employees import employees
from sqlalchemy import func
from sqlalchemy import (
    union_all,
    literal_column,
    asc,
    or_,
    select,
    desc,
    func,
    cast,
    String,
    and_,
    Float,
    extract,
    case,
)
from collections import OrderedDict
from sqlalchemy.orm import aliased
import os, re, math, io, calendar
from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter
import socket, threading, time
from email.message import EmailMessage
from datetime import datetime, timedelta, date
from pytz import timezone
from services.email_service import send_alert_email
from fastapi.encoders import jsonable_encoder

from employees import employees  # import danh sách nhân viên tĩnh

ROLE_MAP = {
    "letan": "Lễ tân",
    "buongphong": "Buồng phòng",
    "quanly": "Quản lý",
    "ktv": "Kỹ thuật viên",
    "baove": "Bảo vệ",
    "boss": "Boss",
    "admin": "Admin",
    "khac": "Khác",
}

def map_role_to_vietnamese(role: Optional[str]) -> str:
    if not role:
        return "Không rõ"
    return ROLE_MAP.get(role.lower(), role.capitalize())

# Định nghĩa múi giờ Việt Nam (UTC+7)
VN_TZ = timezone("Asia/Ho_Chi_Minh")

def get_current_work_shift():
    """
    Xác định ngày và ca làm việc hiện tại.
    - Ca ngày: 07:00 - 18:59
    - Ca đêm: 19:00 - 06:59
    - Thời gian từ 00:00 đến 06:59 được tính là ca đêm của ngày hôm trước.
    """
    now = datetime.now(VN_TZ)
    if now.hour < 7:
        work_date = now.date() - timedelta(days=1)
        shift = "night"
    elif 7 <= now.hour < 19:
        work_date = now.date()
        shift = "day"
    else:  # 19:00 trở đi
        work_date = now.date()
        shift = "night"
    return work_date, shift

def run_daily_absence_check(target_date: Optional[date] = None):
    """
    Chạy kiểm tra và ghi nhận nhân viên vắng mặt.
    Nếu target_date được cung cấp, sẽ chạy cho ngày đó (chạy thủ công).
    Nếu không, sẽ chạy cho ngày hôm trước (dùng cho cron job tự động).
    """
    log_prefix = "thủ công"
    if target_date is None:
        target_date = datetime.now(VN_TZ).date() - timedelta(days=1)
        log_prefix = "tự động"

    logger.info(f"Bắt đầu chạy kiểm tra điểm danh vắng {log_prefix} cho ngày {target_date.strftime('%d/%m/%Y')}")
    # LƯU Ý: Hàm update_missing_attendance_to_db cần được sửa đổi để chấp nhận tham số `target_date`.
    update_missing_attendance_to_db(employees, target_date=target_date)
    logger.info(f"Hoàn tất kiểm tra điểm danh vắng cho ngày {target_date.strftime('%d/%m/%Y')}")

from urllib.parse import parse_qs, urlencode

def clean_query_string(query: str) -> str:
    parsed = parse_qs(query)
    # Loại bỏ các key không mong muốn
    parsed.pop("success", None)
    parsed.pop("action", None)
    return urlencode(parsed, doseq=True)

app = FastAPI()
app.add_middleware(SessionMiddleware, secret_key="binbin-hotel-secret")
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")
templates.env.filters['to_json_serializable'] = jsonable_encoder

@app.get("/ping")
async def ping():
    """
    Endpoint công khai để giữ cho dịch vụ (và database) luôn "thức".
    UptimeRobot hoặc các dịch vụ tương tự có thể gọi endpoint này định kỳ.
    """
    return {"status": "ok"}

BRANCHES = [
    "B1", "B2", "B3",
    "B5", "B6", "B7",
    "B8", "B9", "B10",
    "B11", "B12", "B14",
    "B15", "B16", "B17"
]

from math import radians, cos, sin, sqrt, atan2

branchCoordinates = {
    "B1": [10.727298831515066,106.6967154830272],
    "B2": [10.740600,106.695797],
    "B3": [10.733902,106.708781],
    "B5": [10.73780906347085,106.70517496567874],
    "B6": [10.729986861681768,106.70690372549372],
    "B7": [10.744230207012244,106.6965025304644],
    "B8": [10.741408,106.699883],
    "B9": [10.740970,106.699825],
    "B10": [10.814503,106.670873],
    "B11": [10.77497650247788,106.75134333045331],
    "B12": [10.778874744587053,106.75266727478706],
    "B14": [10.742557513695218,106.69945313180673],
    "B15": [10.775572501574938,106.75167172807936],
    "B16": [10.760347394497392,106.69043939445082],
    "B17": [10.70590976421059, 106.7078826381241], # TODO: Cập nhật tọa độ GPS chính xác cho chi nhánh B17
}

def haversine(lat1, lon1, lat2, lon2):
    R = 6371  # bán kính trái đất km
    dlat = radians(lat2 - lat1)
    dlon = radians(lon2 - lon1)
    a = sin(dlat/2)**2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1-a))
    return R * c

@app.post("/attendance/api/detect-branch")
async def detect_branch(request: Request, db: Session = Depends(get_db)):
    special_roles = ["quanly", "ktv", "boss", "admin"]

    # Lấy user từ session (ưu tiên user, fallback pending_user)
    user_data = request.session.get("user") or request.session.get("pending_user")
    user_in_db = None
    if user_data:
        user_in_db = db.query(User).filter(User.code == user_data["code"]).first()

    # ===============================
    # 1. Role đặc biệt → bỏ qua GPS
    # ===============================
    if user_data and user_data.get("role") in special_roles:
        if user_in_db and user_in_db.branch:
            main_branch = user_in_db.branch
            request.session["active_branch"] = main_branch
            user_in_db.last_active_branch = main_branch
            db.commit()
            return {"branch": main_branch, "distance_km": 0}

        return JSONResponse(
            {"error": "Không thể lấy chi nhánh chính. Vui lòng liên hệ quản trị."},
            status_code=400,
        )

    # ===============================
    # 2. Role thường → dùng GPS
    # ===============================
    try:
        data = await request.json()
    except Exception:
        data = {}

    lat, lng = data.get("lat"), data.get("lng")
    if lat is None or lng is None:
        return JSONResponse(
            {"error": "Bạn vui lòng mở định vị trên điện thoại để lấy vị trí"},
            status_code=400,
        )

    # Tìm chi nhánh trong bán kính 200m
    nearby_branches = []
    for branch, coords in branchCoordinates.items():
        dist = haversine(lat, lng, coords[0], coords[1])
        if dist <= 0.2:  # trong 200m
            nearby_branches.append((branch, dist))

    if not nearby_branches:
        return JSONResponse(
            {"error": "Bạn đang ở quá xa khách sạn (ngoài 200m). Vui lòng điểm danh tại khách sạn."},
            status_code=403,
        )

    # Nếu có nhiều chi nhánh gần → cho frontend chọn
    if len(nearby_branches) > 1:
        choices = [
            {"branch": b, "distance_km": round(d, 3)}
            for b, d in sorted(nearby_branches, key=lambda x: x[1])
        ]
        return {"choices": choices}

    # Nếu chỉ có 1 chi nhánh gần → chọn luôn
    chosen_branch, min_distance = nearby_branches[0]

    request.session["active_branch"] = chosen_branch
    if user_in_db:
        user_in_db.last_active_branch = chosen_branch
        db.commit()

    return {"branch": chosen_branch, "distance_km": round(min_distance, 3)}

@app.post("/attendance/api/select-branch")
async def select_branch(request: Request, db: Session = Depends(get_db)):
    data = await request.json()
    branch = data.get("branch")

    user_data = request.session.get("user") or request.session.get("pending_user")
    if not user_data:
        return JSONResponse({"error": "User chưa đăng nhập"}, status_code=403)

    request.session["active_branch"] = branch

    # Lưu DB
    user_in_db = db.query(User).filter(User.code == user_data["code"]).first()
    if user_in_db:
        user_in_db.last_active_branch = branch
        db.commit()

    return {"branch": branch}

@app.get("/attendance/service", response_class=HTMLResponse)
def attendance_service_ui(request: Request, db: Session = Depends(get_db)):
    user_data = request.session.get("user")
    if not user_data:
        return RedirectResponse("/login", status_code=303)
    # Quản lý và KTV không có chức năng chấm dịch vụ
    if user_data.get("role") in ["quanly", "ktv", "admin", "boss"]:
        return RedirectResponse("/choose-function", status_code=303)

    checker_user = db.query(User).filter(User.code == user_data["code"]).first()
    active_branch = request.session.get("active_branch")
    # Nếu trong session không có, thử lấy từ DB (lần đăng nhập trước)
    if not active_branch and checker_user and hasattr(checker_user, 'last_active_branch') and checker_user.last_active_branch:
        active_branch = checker_user.last_active_branch
        request.session["active_branch"] = active_branch # Lưu lại vào session cho lần tải trang sau trong cùng phiên
    # Nếu vẫn không có, dùng chi nhánh mặc định của user
    if not active_branch:
        active_branch = user_data.get("branch", "")
    csrf_token = get_csrf_token(request)

    initial_employees = []
    # Lấy danh sách nhân viên BP đã được điểm danh lần cuối từ DB của lễ tân
    if checker_user and checker_user.last_checked_in_bp:
        service_checkin_codes = checker_user.last_checked_in_bp

        if service_checkin_codes:
            # Lấy thông tin chi tiết của các nhân viên đó
            employees_from_db = db.query(User).filter(User.code.in_(service_checkin_codes)).all()
            # Sắp xếp lại theo đúng thứ tự đã điểm danh
            emp_map = {emp.code: emp for emp in employees_from_db}
            sorted_employees = [emp_map[code] for code in service_checkin_codes if code in emp_map]
            
            initial_employees = [
                {"code": emp.code, "name": emp.name, "branch": emp.branch, "so_phong": "", "so_luong": "", "dich_vu": "", "ghi_chu": ""}
                for emp in sorted_employees
            ]

    response = templates.TemplateResponse("attendance_service.html", {
        "request": request,
        "branch_id": active_branch,
        "csrf_token": csrf_token,
        "user": user_data,
        "initial_employees": initial_employees,
    })
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

def _get_filtered_tasks_query(
    db: Session,
    user_data: dict,
    chi_nhanh: str = "",
    search: str = "",
    trang_thai: str = "",
    han_hoan_thanh: str = ""
):
    """
    Hàm helper để xây dựng và trả về câu truy vấn SQLAlchemy cho các công việc
    dựa trên các bộ lọc được cung cấp.
    """
    role = user_data.get("role")

    tasks_query = db.query(Task)

    # Loại bỏ công việc đã xoá cho các vai trò không phải quản lý cấp cao
    if role not in ["quanly", "admin", "boss"]:
        tasks_query = tasks_query.filter(Task.trang_thai != "Đã xoá")

    # Lọc theo chi nhánh (nếu có).
    # `chi_nhanh` ở đây đã được xác định một cách chính xác ở hàm `home`.
    if chi_nhanh:
        tasks_query = tasks_query.filter(Task.chi_nhanh == chi_nhanh)

    # Lọc theo từ khóa
    if search:
        clean_search = re.sub(r'\s+', ' ', search).strip()
        search_pattern = f"%{clean_search}%"
        tasks_query = tasks_query.filter(
            or_(
                Task.chi_nhanh.ilike(search_pattern),
                Task.phong.ilike(search_pattern),
                Task.mo_ta.ilike(search_pattern),
                Task.trang_thai.ilike(search_pattern),
                Task.nguoi_tao.ilike(search_pattern),
                Task.nguoi_thuc_hien.ilike(search_pattern),
                Task.ghi_chu.ilike(search_pattern)
            )
        )

    # Lọc theo trạng thái
    if trang_thai:
        tasks_query = tasks_query.filter(Task.trang_thai == trang_thai)

    # Lọc theo hạn hoàn thành
    if han_hoan_thanh:
        try:
            han_date = datetime.strptime(han_hoan_thanh, "%Y-%m-%d").date()
            tasks_query = tasks_query.filter(func.date(Task.han_hoan_thanh) == han_date)
        except (ValueError, TypeError):
            pass  # Bỏ qua nếu định dạng ngày không hợp lệ

    return tasks_query

@app.get("/", response_class=HTMLResponse)
def root(request: Request):
    """Route gốc, chuyển hướng người dùng dựa trên trạng thái đăng nhập."""
    if request.session.get("user"):
        return RedirectResponse("/login", status_code=303) 
    return RedirectResponse("/login", status_code=303)

# --- Sử dụng middleware này ở các route yêu cầu đăng nhập ---
@app.get("/choose-function", response_class=HTMLResponse)
async def choose_function(request: Request):
    if not require_checked_in_user(request): # This check also ensures user is in session
        return RedirectResponse("/login", status_code=303)

    # Nếu có flag after_checkin thì xóa để tránh dùng lại
    if request.session.get("after_checkin") == "choose_function":
        request.session.pop("after_checkin", None)

    response = templates.TemplateResponse(
        "choose_function.html", {"request": request, "user": request.session.get("user")}
    )
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

@app.get("/attendance/results", response_class=HTMLResponse)
def view_attendance_results(request: Request, db: Session = Depends(get_db)):
    """
    Route để hiển thị trang xem kết quả điểm danh.
    """
    if not require_checked_in_user(request):
        return RedirectResponse("/login", status_code=303)

    user_data = request.session.get("user")
    
    # Tạo một bản sao của ROLE_MAP và loại bỏ vai trò 'khac' để không hiển thị trong bộ lọc
    roles_for_filter = {k: v for k, v in ROLE_MAP.items() if k != 'khac'}

    return templates.TemplateResponse("attendance_results.html", {
        "request": request,
        "user": user_data,
        "branches": BRANCHES,
        "roles": roles_for_filter,
        "dashboard_stats": None
    })

@app.get("/attendance/calendar-view", response_class=HTMLResponse)
def view_attendance_calendar(
    request: Request,
    db: Session = Depends(get_db),
    chi_nhanh: Optional[str] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
):
    user_data = request.session.get("user")
    # Cấp quyền cho Lễ tân
    if not user_data or user_data.get("role") not in ["admin", "boss", "quanly", "letan", "ktv"]:
        return RedirectResponse("/choose-function", status_code=303)

    # Tạo danh sách chi nhánh để hiển thị trong bộ lọc
    display_branches = BRANCHES.copy()
    if user_data.get("role") in ["admin", "boss"]:
        display_branches.extend(["KTV", "Quản lý", "LTTC", "BPTC"])

    # Nếu chưa có chi nhánh được chọn từ filter, đặt giá trị mặc định theo vai trò
    if not chi_nhanh:
        user_role = user_data.get("role")
        user_code = user_data.get("code", "")

        if user_role == "ktv":
            # KTVs default to the "KTV" role-based view
            chi_nhanh = "KTV"
        elif user_role == "quanly":
            # Managers default to the "Quản lý" role-based view
            chi_nhanh = "Quản lý"
        elif user_role == "letan" and "LTTC" in user_code.upper():
            # LTTC users default to the "LTTC" code-based view
            chi_nhanh = "LTTC"
        elif user_role == "letan": # Lễ tân thường: ưu tiên GPS/last active
            user_from_db = db.query(User).filter(User.code == user_data["code"]).first()
            active_branch = (
                request.session.get("active_branch")
                or (user_from_db.last_active_branch if user_from_db and hasattr(user_from_db, 'last_active_branch') else None)
                or user_data.get("branch")
            )
            chi_nhanh = active_branch
        elif user_role in ["admin", "boss"]:
            chi_nhanh = "B1"

    now = datetime.now(VN_TZ)
    current_month = month if month else now.month
    current_year = year if year else now.year

    # Mở rộng khoảng thời gian query để lấy ca đêm của ngày cuối tháng trước và ca đêm của ngày cuối tháng này
    start_query_date = date(current_year, current_month, 1) - timedelta(days=1)
    end_query_date = date(current_year, current_month, calendar.monthrange(current_year, current_month)[1]) + timedelta(days=1)

    _, num_days = calendar.monthrange(current_year, current_month)
    
    employee_data = defaultdict(lambda: {
        "name": "",
        "role": "",
        "role_key": "",
        "main_branch": "",
        "is_cross_branch_month": False, # Sẽ được cập nhật sau
        "worked_away_from_main_branch": False, # Cờ mới để đánh dấu *
        "daily_work": defaultdict(lambda: {"work_units": 0, "is_overtime": False, "work_branch": "", "services": []})
    })

    if chi_nhanh:
        # --- BƯỚC 1: Lấy danh sách nhân viên cơ sở thuộc về chi nhánh/vai trò đang xem ---
        # Điều này đảm bảo tất cả nhân viên chính thức sẽ luôn xuất hiện trong lịch,
        # ngay cả khi họ không có ngày công nào trong tháng.
        base_employee_query = db.query(User.code, User.name, User.role, User.branch)
        
        role_map_filter = {"KTV": "ktv", "Quản lý": "quanly"}
        code_prefix_filter = {"LTTC": "LTTC", "BPTC": "BPTC"}

        if chi_nhanh in role_map_filter:
            base_employee_query = base_employee_query.filter(User.role == role_map_filter[chi_nhanh])
        elif chi_nhanh in code_prefix_filter:
            base_employee_query = base_employee_query.filter(User.code.startswith(code_prefix_filter[chi_nhanh]))
        elif chi_nhanh in BRANCHES:
            base_employee_query = base_employee_query.filter(User.branch == chi_nhanh)
        else:
            base_employee_query = None # Không có bộ lọc hợp lệ

        if base_employee_query:
            base_employees = base_employee_query.all()
            for emp in base_employees:
                emp_code = emp.code
                if emp_code not in employee_data:
                    employee_data[emp_code]["name"] = emp.name
                    employee_data[emp_code]["main_branch"] = emp.branch

                    # Xác định role_key để sắp xếp, ưu tiên LTTC/BPTC
                    emp_role_key = emp.role or "khac"
                    if "LTTC" in emp_code.upper():
                        emp_role_key = "lttc"
                    elif "BPTC" in emp_code.upper():
                        emp_role_key = "bptc"
                    
                    employee_data[emp_code]["role_key"] = emp_role_key
                    employee_data[emp_code]["role"] = map_role_to_vietnamese(emp.role)


        # --- Xây dựng bộ lọc dựa trên lựa chọn ---
        att_location_filter = None
        svc_location_filter = None

        role_map_filter = {"KTV": "ktv", "Quản lý": "quanly"}
        code_prefix_filter = {"LTTC": "LTTC", "BPTC": "BPTC"}

        if chi_nhanh in role_map_filter: # Lọc bản ghi theo vai trò
            # Lọc theo vai trò nếu chọn KTV hoặc Quản lý
            role_to_filter = role_map_filter[chi_nhanh]
            att_location_filter = (User.role == role_to_filter)
            svc_location_filter = (User.role == role_to_filter)
        elif chi_nhanh in code_prefix_filter:
            # Lọc theo mã nhân viên nếu chọn LTTC/BPTC
            prefix_to_filter = code_prefix_filter[chi_nhanh]
            att_location_filter = (User.code.startswith(prefix_to_filter))
            svc_location_filter = (User.code.startswith(prefix_to_filter))
        elif chi_nhanh in BRANCHES: # Lọc bản ghi theo chi nhánh
            # Chỉ lấy các bản ghi chấm công/dịch vụ được thực hiện TẠI chi nhánh đang xem.
            # Danh sách nhân viên đầy đủ của chi nhánh đã được lấy ở Bước 1,
            # nên ở đây ta không cần lấy tất cả các bản ghi của họ ở những nơi khác nữa.
            att_location_filter = (AttendanceRecord.chi_nhanh_lam == chi_nhanh)
            svc_location_filter = (ServiceRecord.chi_nhanh_lam == chi_nhanh)

        
        # Chỉ thực hiện query nếu có bộ lọc hợp lệ
        if att_location_filter is not None:
            # Query cho điểm danh
            att_q = select(
                literal_column("'attendance'").label("type"),
                AttendanceRecord.ma_nv, AttendanceRecord.ten_nv, User.role, User.branch.label("main_branch"),
                AttendanceRecord.ngay_diem_danh.label("date_col"),
                AttendanceRecord.gio_diem_danh.label("time_col"),
                AttendanceRecord.so_cong_nv.label("value"),
                AttendanceRecord.la_tang_ca,
                AttendanceRecord.chi_nhanh_lam.label("work_branch"),
                literal_column("''").label("dich_vu")
            ).join(
                User, User.code == AttendanceRecord.ma_nv, isouter=True
            ).filter(
                AttendanceRecord.ngay_diem_danh.between(start_query_date, end_query_date),
                att_location_filter
            )

            # Query cho dịch vụ
            svc_q = select(
                literal_column("'service'").label("type"),
                ServiceRecord.ma_nv, ServiceRecord.ten_nv, User.role, User.branch.label("main_branch"),
                ServiceRecord.ngay_cham.label("date_col"),
                ServiceRecord.gio_cham.label("time_col"),
                cast(ServiceRecord.so_luong, Float).label("value"),
                ServiceRecord.la_tang_ca,
                ServiceRecord.chi_nhanh_lam.label("work_branch"),
                ServiceRecord.dich_vu
            ).join(
                User, User.code == ServiceRecord.ma_nv, isouter=True
            ).filter(
                ServiceRecord.ngay_cham.between(start_query_date, end_query_date),
                svc_location_filter
            )

            # Gộp 2 query
            combined_query = union_all(att_q, svc_q).alias("combined")
            records = db.execute(select(combined_query).order_by(combined_query.c.ten_nv, combined_query.c.date_col)).all()

        # Process records into the desired structure
        for rec in records:
            emp_code = rec.ma_nv

            # Xác định ngày làm việc (ca đêm < 7h sáng tính cho ngày hôm trước)
            work_date = rec.date_col - timedelta(days=1) if rec.time_col.hour < 7 else rec.date_col

            # Bỏ qua các bản ghi không thuộc tháng đang xem
            if work_date.month != current_month or work_date.year != current_year:
                continue

            day_of_month = work_date.day

            if not employee_data[emp_code]["name"]:
                employee_data[emp_code]["name"] = rec.ten_nv
                
                # Xác định role_key để sắp xếp, ưu tiên LTTC/BPTC
                emp_role_key = rec.role or "khac"
                if "LTTC" in emp_code.upper():
                    emp_role_key = "lttc"
                elif "BPTC" in emp_code.upper():
                    emp_role_key = "bptc"
                employee_data[emp_code]["role_key"] = emp_role_key
                employee_data[emp_code]["role"] = map_role_to_vietnamese(rec.role)
                employee_data[emp_code]["main_branch"] = rec.main_branch

            if not employee_data[emp_code]["is_cross_branch_month"] and rec.main_branch != chi_nhanh:
                employee_data[emp_code]["is_cross_branch_month"] = True

            # Logic mới: Đánh dấu nếu nhân viên từng làm khác chi nhánh chính trong tháng
            if rec.work_branch and rec.main_branch and rec.work_branch != rec.main_branch:
                employee_data[emp_code]["worked_away_from_main_branch"] = True

            if rec.type == 'attendance':
                daily_work_entry = employee_data[emp_code]["daily_work"][day_of_month]
                daily_work_entry["work_units"] += rec.value or 0
                if rec.la_tang_ca:
                    daily_work_entry["is_overtime"] = True
                    daily_work_entry["work_branch"] = rec.work_branch
            elif rec.type == 'service':
                daily_work_entry = employee_data[emp_code]["daily_work"][day_of_month]
                if rec.la_tang_ca:
                    daily_work_entry["is_overtime"] = True
                    daily_work_entry["work_branch"] = rec.work_branch
                # Tổng hợp dịch vụ theo ngày
                service_summary = daily_work_entry.setdefault("service_summary", defaultdict(int))
                service_summary[rec.dich_vu] += int(rec.value or 0)

        # Chuyển đổi service_summary thành list string để dễ render
        for emp_code in employee_data:
            for day in employee_data[emp_code]["daily_work"]:
                if "service_summary" in employee_data[emp_code]["daily_work"][day]:
                    summary = employee_data[emp_code]["daily_work"][day].pop("service_summary")
                    employee_data[emp_code]["daily_work"][day]["services"] = [f"{k}: {v}" for k, v in summary.items()]
    
        # --- TÍNH TOÁN THỐNG KÊ DASHBOARD CHO NHÂN VIÊN CỦA CHI NHÁNH ĐANG XEM ---
        for emp_code, emp_details in employee_data.items():
            # Chỉ tính cho nhân viên có chi nhánh chính là chi nhánh đang xem
            is_main_employee_of_view = (
                emp_details.get("main_branch") == chi_nhanh
                or (chi_nhanh == "KTV" and emp_details.get("role_key") == "ktv")
                or (chi_nhanh == "Quản lý" and emp_details.get("role_key") == "quanly")
                or (chi_nhanh == "LTTC" and emp_details.get("role_key") == "lttc")
                or (chi_nhanh == "BPTC" and emp_details.get("role_key") == "bptc")
            )

            if is_main_employee_of_view:
                # --- TÍNH TOÁN DASHBOARD ---
                # 1. Lấy tất cả bản ghi điểm danh có khả năng ảnh hưởng đến tháng đang xem
                # (bao gồm cả ca đêm của ngày cuối tháng trước và ca đêm của ngày cuối tháng này)
                start_query_date = date(current_year, current_month, 1)
                end_query_date = date(current_year, current_month, num_days) + timedelta(days=1)
                
                all_atts_raw = db.query(
                    AttendanceRecord.ngay_diem_danh, AttendanceRecord.gio_diem_danh,
                    AttendanceRecord.so_cong_nv, AttendanceRecord.la_tang_ca,
                    AttendanceRecord.chi_nhanh_lam
                ).filter(
                    AttendanceRecord.ma_nv == emp_code,
                    AttendanceRecord.ngay_diem_danh.between(start_query_date, end_query_date)
                ).all()

                # Helper để xác định ngày làm việc (ca đêm < 7h sáng tính cho ngày hôm trước)
                def get_work_day(att_date, att_time):
                    return att_date - timedelta(days=1) if att_time.hour < 7 else att_date

                # Gắn "work_day" vào mỗi bản ghi và lọc lại theo tháng đang xem
                all_atts = [
                    {**att._asdict(), "work_day": get_work_day(att.ngay_diem_danh, att.gio_diem_danh)}
                    for att in all_atts_raw
                ]
                all_atts = [
                    att for att in all_atts 
                    if att["work_day"].month == current_month and att["work_day"].year == current_year
                ]

                # 2. Xử lý dữ liệu điểm danh dựa trên "work_day"
                tong_so_cong = 0.0
                work_days_set = set()
                overtime_work_days_set = set()
                daily_work_units = defaultdict(float)

                for att in all_atts:
                    work_day = att['work_day']
                    so_cong = att['so_cong_nv'] or 0
                    tong_so_cong += so_cong
                    if so_cong > 0:
                        work_days_set.add(work_day)
                    daily_work_units[work_day] += so_cong
                    if att['la_tang_ca']:
                        overtime_work_days_set.add(work_day)

                # Xác định ngày tăng ca dựa trên tổng công > 1
                for day, total_units in daily_work_units.items():
                    if total_units > 1:
                        overtime_work_days_set.add(day)

                # Lấy chi tiết tăng ca
                overtime_details = []
                main_branch = emp_details.get("main_branch")

                # Xác định những ngày làm việc có chấm công ở chi nhánh khác (với số công > 0)
                other_branch_work_days = {
                    att['work_day']
                    for att in all_atts
                    if main_branch and att['chi_nhanh_lam'] != main_branch and (att['so_cong_nv'] or 0) > 0
                }

                # Set để đảm bảo mỗi ngày chỉ xử lý 1 lần cho trường hợp >1 công
                processed_main_branch_overtime_days = set()

                # Lặp qua tất cả các bản ghi để xây dựng chi tiết
                for att in all_atts:
                    work_day = att['work_day']

                    # Bỏ qua nếu không phải là ngày tăng ca
                    if work_day not in overtime_work_days_set:
                        continue

                    # Ưu tiên 1: Tăng ca do đi chi nhánh khác
                    if work_day in other_branch_work_days:
                        # Chỉ thêm các bản ghi ở chi nhánh khác (có công)
                        if main_branch and att['chi_nhanh_lam'] != main_branch and (att['so_cong_nv'] or 0) > 0:
                            overtime_details.append({ "date": att['ngay_diem_danh'].strftime('%d/%m/%Y'), "time": att['gio_diem_danh'].strftime('%H:%M'), "branch": att['chi_nhanh_lam'], "work_units": att['so_cong_nv'] })
                    # Trường hợp 2: Tăng ca do làm >1 công (và chỉ làm tại chi nhánh chính)
                    elif daily_work_units.get(work_day, 0) > 1:
                        if work_day not in processed_main_branch_overtime_days:
                            # Chỉ hiển thị 1 dòng tóm tắt cho ngày này
                            overtime_details.append({ "date": work_day.strftime('%d/%m/%Y'), "time": "Nhiều ca", "branch": main_branch, "work_units": f"{daily_work_units.get(work_day, 0):.1f}" })
                            processed_main_branch_overtime_days.add(work_day)

                # 3. Lấy tất cả bản ghi dịch vụ trong tháng
                all_services = db.query(
                    ServiceRecord
                ).filter(
                    ServiceRecord.ma_nv == emp_code,
                    extract('month', ServiceRecord.ngay_cham) == current_month,
                    extract('year', ServiceRecord.ngay_cham) == current_year
                ).all()

                # 4. Tổng hợp kết quả
                so_ngay_lam = len(work_days_set)
                so_ngay_tang_ca = len(overtime_work_days_set)

                # --- LOGIC MỚI CHO SỐ NGÀY NGHỈ ---
                is_current_month_view = (current_year == now.year and current_month == now.month)
                
                if is_current_month_view:
                    # Đối với tháng hiện tại, số ngày nghỉ được tính từ đầu tháng đến ngày hôm nay.
                    days_passed = now.day
                    # Lọc ra những ngày đã làm việc tính đến hôm nay.
                    worked_days_so_far = {d for d in work_days_set if d <= now.date()}
                    so_ngay_nghi = days_passed - len(worked_days_so_far)
                else:
                    # Đối với các tháng trong quá khứ, tính như cũ.
                    so_ngay_nghi = num_days - so_ngay_lam
                so_ngay_nghi = max(0, so_ngay_nghi)
                laundry_details = []
                ironing_details = []
                tong_dich_vu_giat = 0
                tong_dich_vu_ui = 0

                for svc in all_services:
                    try:
                        quantity = int(svc.so_luong)
                    except (ValueError, TypeError):
                        quantity = 0
                    
                    detail = {
                        "date": svc.ngay_cham.strftime('%d/%m/%Y'), "time": svc.gio_cham.strftime('%H:%M'),
                        "branch": svc.chi_nhanh_lam, "room": svc.so_phong, "quantity": svc.so_luong
                    }

                    if svc.dich_vu == 'Giặt':
                        tong_dich_vu_giat += quantity
                        laundry_details.append(detail)
                    elif svc.dich_vu == 'Ủi':
                        tong_dich_vu_ui += quantity
                        ironing_details.append(detail)

                emp_details["dashboard_stats"] = {
                    "so_ngay_lam": so_ngay_lam,
                    "so_ngay_nghi": so_ngay_nghi,
                    "so_ngay_tang_ca": so_ngay_tang_ca,
                    "tong_so_cong": tong_so_cong,
                    "tong_dich_vu_giat": tong_dich_vu_giat,
                    "tong_dich_vu_ui": tong_dich_vu_ui,
                    "overtime_details": sorted(overtime_details, key=lambda x: datetime.strptime(x['date'], '%d/%m/%Y')),
                    "laundry_details": sorted(laundry_details, key=lambda x: datetime.strptime(x['date'], '%d/%m/%Y')),
                    "ironing_details": sorted(ironing_details, key=lambda x: datetime.strptime(x['date'], '%d/%m/%Y')),
                }
    
    # Sắp xếp nhân viên theo: Chức vụ > Tăng ca (làm khác CN) > Tên
    role_priority = {
        "letan": 0,
        "lttc": 0,      # Sắp xếp cùng Lễ tân
        "buongphong": 1,
        "bptc": 1,      # Sắp xếp cùng Buồng phòng
        "baove": 2,
        "ktv": 3,
        "quanly": 4,
    }
    
    # Chuyển dict thành list để sắp xếp
    employee_list_to_sort = list(employee_data.items())

    # Sắp xếp list
    sorted_employee_list = sorted(
        employee_list_to_sort,
        key=lambda item: (
            role_priority.get(item[1].get("role_key", "khac"), 99),
            item[1].get("is_cross_branch_month", False),
            item[1].get("name", "")
        )
    )
    sorted_employee_data = OrderedDict(sorted_employee_list)

    # Chuẩn bị dữ liệu chi tiết cho JavaScript
    employee_data_for_js = {}
    for code, data in sorted_employee_data.items():
        if "dashboard_stats" in data:
            employee_data_for_js[code] = {
                "dashboard_stats": data["dashboard_stats"]
            }

    return templates.TemplateResponse("attendance_calendar_view.html", {
        "request": request,
        "user": user_data,
        "branches": display_branches,
        "selected_branch": chi_nhanh,
        "selected_month": current_month,
        "selected_year": current_year,
        "num_days": num_days,
        "employee_data": sorted_employee_data,
        "employee_data_for_js": employee_data_for_js,
        "current_day": now.day if now.month == current_month and now.year == current_year else None,
    })

@app.get("/login", response_class=HTMLResponse)
def login_form(request: Request):
    # Nếu người dùng đã đăng nhập VÀ đã điểm danh QR thành công hôm nay thì chuyển về trang chọn chức năng
    user = request.session.get("user")
    if user:
        today = date.today()
        db = SessionLocal()
        try:
            log = db.query(AttendanceLog).filter_by(user_code=user["code"], date=today).first()
            if log and log.checked_in:
                    return RedirectResponse(url="/choose-function", status_code=303)
        finally:
            db.close()
    error = request.query_params.get("error", "")
    role = request.query_params.get("role", "")
    response = templates.TemplateResponse("login.html", {
        "request": request,
        "error": error,
        "role": role
    })
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

def _get_log_shift_for_user(user_role: str, current_shift: str) -> Optional[str]:
    """Xác định ca làm việc để ghi log, trả về None cho các role đặc biệt."""
    return None if user_role in ["ktv", "quanly"] else current_shift

@app.post("/login", response_class=HTMLResponse)
def login_submit(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    request.session.clear()
    allowed_roles = ["letan", "quanly", "ktv", "admin", "boss"]

    # Tìm user
    user = db.query(User).filter(
        User.code == username,
        User.password == password,
        User.role.in_(allowed_roles)
    ).first()

    if not user:
        # Nếu login sai → đoán role để hiển thị UI đúng
        guessed_role = ""
        if username.lower().startswith("b") and "lt" in username.lower():
            guessed_role = "letan"
        elif username.lower().startswith("ktv"):
            guessed_role = "ktv"
        elif username.lower() in ["ql", "admin"]:
            guessed_role = "quanly"

        query = urlencode({
            "error": "Mã nhân viên hoặc mật khẩu sai",
            "role": guessed_role
        })
        return RedirectResponse(f"/login?{query}", status_code=303)

    # ✅ Boss & Admin: vào thẳng hệ thống, không cần điểm danh
    if user.role in ["boss", "admin"]:
        request.session["user"] = {
            "code": user.code,
            "role": user.role,
            "branch": user.branch,
            "name": user.name
        }
        request.session["after_checkin"] = "choose_function"
        return RedirectResponse("/choose-function", status_code=303)

    # ✅ Các role khác: kiểm tra log điểm danh
    work_date, shift = get_current_work_shift()
    shift_value = _get_log_shift_for_user(user.role, shift)

    # Query log theo shift_value
    log = db.query(AttendanceLog).filter_by(
        user_code=user.code,
        date=work_date,
        shift=shift_value
    ).first()

    # Nếu đã check-in thì vào thẳng hệ thống
    if log and log.checked_in:
        request.session["user"] = {
            "code": user.code,
            "role": user.role,
            "branch": user.branch,
            "name": user.name
        }
        request.session.pop("pending_user", None)
        return RedirectResponse("/choose-function", status_code=303)

    # Nếu chưa check-in → phân luồng mobile / desktop
    user_agent = request.headers.get("user-agent", "").lower()
    is_mobile = any(k in user_agent for k in ["mobi", "android", "iphone", "ipad"])

    if is_mobile:
        if not log:
            token = secrets.token_urlsafe(24)
            log = AttendanceLog(
                user_code=user.code,
                date=work_date,
                shift=shift_value,
                token=token,
                checked_in=False
            )
            db.add(log)
            db.commit()
        token = log.token

        request.session["pending_user"] = {
            "code": user.code,
            "role": user.role,
            "branch": user.branch,
            "name": user.name,
        }
        request.session["qr_token"] = token
        return RedirectResponse("/attendance/ui", status_code=303)

    else:
        # Desktop
        if log and not log.checked_in:
            token = log.token
        else:
            token = secrets.token_urlsafe(24)
            log = AttendanceLog(
                user_code=user.code,
                date=work_date,
                shift=shift_value,
                token=token,
                checked_in=False
            )
            db.add(log)
            db.commit()

        request.session["pending_user"] = {
            "code": user.code,
            "role": user.role,
            "branch": user.branch,
            "name": user.name,
        }
        request.session["qr_token"] = token
        return RedirectResponse("/show_qr", status_code=303)

# --- Middleware kiểm tra trạng thái điểm danh QR ---
from datetime import date

def require_checked_in_user(request: Request):
    user = request.session.get("user")
    if not user:
        return False

    # ✅ Admin và Boss luôn được truy cập nếu đã đăng nhập
    if user.get("role") in ["admin", "boss"]:
        return True

    work_date, _ = get_current_work_shift()
    db = SessionLocal()
    try:
        # Kiểm tra xem có bất kỳ log nào đã check-in trong ngày làm việc hiện tại không
        # (ca ngày hoặc ca đêm).
        log = db.query(AttendanceLog).filter(
            AttendanceLog.user_code == user["code"],
            AttendanceLog.date == work_date,
            AttendanceLog.checked_in == True
        ).first()

        # Cho phép vào nếu có log checked_in trong DB hoặc vừa quét QR xong
        if (log and log.checked_in) or request.session.get("after_checkin") == "choose_function":
            return True
    finally:
        db.close()

    return False

# --- CSRF Token Management ---
def generate_csrf_token():
    return secrets.token_urlsafe(32)

def get_csrf_token(request: Request):
    token = request.session.get("csrf_token")
    if not token:
        token = generate_csrf_token()
        request.session["csrf_token"] = token
    return token

def validate_csrf(request: Request):
    token = request.headers.get("X-CSRF-Token") or request.query_params.get("csrf_token")
    session_token = request.session.get("csrf_token")
    if not session_token or token != session_token:
        raise HTTPException(status_code=403, detail="CSRF token không hợp lệ")

# --- Attendance UI ---
@app.get("/attendance/ui", response_class=HTMLResponse)
def attendance_ui(request: Request):
    user_data = request.session.get("user") or request.session.get("pending_user")
    if not user_data:
        return RedirectResponse("/login", status_code=303)
    active_branch = request.session.get("active_branch") or user_data.get("branch", "")
    csrf_token = get_csrf_token(request)
    # Truyền mã nhân viên đăng nhập cho frontend để luôn hiển thị trong danh sách điểm danh
    response = templates.TemplateResponse("attendance.html", {
        "request": request,
        "branch_id": active_branch,
        "csrf_token": csrf_token,
        # "branches": BRANCHES,
        "user": user_data,
        "login_code": user_data.get("code", ""),  # thêm dòng này
    })
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

# --- Get CSRF token ---
@app.get("/attendance/csrf-token")
def attendance_csrf_token(request: Request):
    token = get_csrf_token(request)
    return {"csrf_token": token}

@app.get("/attendance/api/employees/by-branch/{branch_id}", response_class=JSONResponse)
def get_employees_by_branch(branch_id: str, db: Session = Depends(get_db), request: Request = None):
    try:
        user = request.session.get("user") if request else None
        now_hour = datetime.now(VN_TZ).hour
        current_shift = "CS" if 7 <= now_hour < 19 else "CT"

        def match_shift(emp_code: str):
            emp_code = emp_code.upper()
            if "CS" in emp_code and current_shift != "CS":
                return False
            if "CT" in emp_code and current_shift != "CT":
                return False
            return True

        employees = []

        if user and user.get("role") == "letan":
            # ✅ Luôn thêm chính lễ tân đang đăng nhập
            lt_self = db.query(User).filter(
                User.code == user.get("code"),
                User.branch == branch_id
            ).all()
            # Các bộ phận khác cùng chi nhánh (bỏ quản lý, ktv, lễ tân khác)
            others = db.query(User).filter(
                User.branch == branch_id,
                ~User.role.in_(["quanly", "ktv", "letan"])
            ).all()
            others = [emp for emp in others if match_shift(emp.code)]
            employees = sorted(lt_self + others, key=lambda e: e.name)

        elif user and user.get("role") in ["quanly", "ktv"]:
            # ✅ Quản lý và KTV chỉ thấy chính họ (bỏ lọc chi nhánh, bỏ shift)
            employees = db.query(User).filter(
                User.code == user.get("code")
            ).all()

        elif user and user.get("role") in ["admin", "boss"]:
            # ✅ Admin và Boss thấy tất cả nhân viên của chi nhánh, không lọc shift
            employees = db.query(User).filter(User.branch == branch_id).order_by(User.name).all()

        else:
            # ✅ Logic chung cho các role khác
            employees = db.query(User).filter(
                User.branch == branch_id,
                ~User.role.in_(["quanly", "ktv", "admin", "boss"])
            ).all()
            employees = [emp for emp in employees if match_shift(emp.code)]
            employees.sort(key=lambda e: e.name)

        employee_list = [
            {"code": emp.code, "name": emp.name, "department": emp.role, "branch": emp.branch}
            for emp in employees
        ]
        return JSONResponse(content=employee_list)

    except Exception as e:
        return JSONResponse(status_code=500, content={"detail": f"Lỗi server: {str(e)}"})

# --- Employee search ---
@app.get("/attendance/api/employees/search", response_class=JSONResponse)
def search_employees(
    q: str = "",
    request: Request = None,
    branch_id: Optional[str] = None,
    only_bp: bool = False,
    loginCode: Optional[str] = None,
    context: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    API tìm kiếm nhân viên theo mã hoặc tên.
    - Nếu branch_id được cung cấp, chỉ tìm trong chi nhánh đó (ngoại trừ khi only_bp=True).
    - Nếu only_bp=True thì chỉ trả về nhân viên buồng phòng (mã chứa 'BP') từ TẤT CẢ các chi nhánh.
    - Mặc định loại bỏ role lễ tân, ngoại trừ lễ tân đang đăng nhập (loginCode) trong các context khác.
    - Thêm context='results_filter' để chỉ gợi ý nhân viên mà user đã chấm công.
    - Giới hạn 20 kết quả.
    """
    if not q:
        if not only_bp:
            return JSONResponse(content=[], status_code=400)
        search_pattern = "%"
    else:
        search_pattern = f"%{q}%"

    user = request.session.get("user") if request else None

    # --- LOGIC MỚI: Lọc theo ngữ cảnh trang kết quả ---
    if user and user.get("role") not in ["admin", "boss"] and context == "results_filter":
        checker_code = user.get("code")

        # Lấy mã các nhân viên mà user này đã tạo bản ghi cho
        att_employee_codes_q = db.query(AttendanceRecord.ma_nv).filter(AttendanceRecord.nguoi_diem_danh == checker_code).distinct()
        svc_employee_codes_q = db.query(ServiceRecord.ma_nv).filter(ServiceRecord.nguoi_cham == checker_code).distinct()

        related_codes = {row[0] for row in att_employee_codes_q.all()}
        related_codes.update({row[0] for row in svc_employee_codes_q.all()})

        # Luôn bao gồm chính mình trong danh sách tìm kiếm (trường hợp tự chấm công)
        related_codes.add(checker_code)

        if not related_codes:
             return JSONResponse(content=[])

        # Xây dựng query dựa trên các mã đã thu thập
        query = db.query(User).filter(
            User.code.in_(list(related_codes)),
            or_(
                User.code.ilike(search_pattern),
                User.name.ilike(search_pattern)
            )
        )
        employees = query.limit(50).all()
    elif context == 'reporter_search':
        # --- LOGIC MỚI: Dùng cho tìm kiếm người báo cáo trong module Đồ thất lạc ---
        # Trả về tất cả nhân viên trừ admin và boss
        query = db.query(User).filter(
            ~User.role.in_(['admin', 'boss']),
            or_(
                User.code.ilike(search_pattern),
                User.name.ilike(search_pattern)
            )
        )
        employees = query.limit(20).all()
        employee_list = [{"code": emp.code, "name": emp.name} for emp in employees]
        return JSONResponse(content=employee_list)
    elif context == 'all_users_search':
        # --- LOGIC MỚI: Dùng cho tìm kiếm tất cả người dùng (bao gồm admin/boss) ---
        # Được sử dụng cho ô "Người thanh lý" trong module Đồ thất lạc.
        query = db.query(User).filter(
            or_(
                User.code.ilike(search_pattern),
                User.name.ilike(search_pattern)
            )
        )
        employees = query.limit(20).all()
        employee_list = [{"code": emp.code, "name": emp.name} for emp in employees]
        return JSONResponse(content=employee_list)

    else:
        # --- LOGIC CŨ: Dùng cho các trường hợp khác (trang điểm danh, admin/boss, etc.) ---
        query = db.query(User).filter(
            or_(
                User.code.ilike(search_pattern),
                User.name.ilike(search_pattern)
            )
        )

        if branch_id and not only_bp:
            query = query.filter(User.branch == branch_id)

        employees = query.limit(50).all()

        if only_bp:
            employees = [emp for emp in employees if "BP" in (emp.code or "").upper()]

        is_admin_or_boss = user and user.get("role") in ["admin", "boss"]

        if not is_admin_or_boss:
            filtered = []
            for emp in employees:
                if (emp.role or "").lower() == "letan":
                    if loginCode and emp.code == loginCode:
                        filtered.append(emp)
                else:
                    filtered.append(emp)
            employees = filtered

    employee_list = [
        {"code": emp.code, "name": emp.name, "department": emp.role, "branch": emp.branch}
        for emp in employees[:20]
    ]
    return JSONResponse(content=employee_list)

from sqlalchemy import case

def _get_filtered_tasks_query(
    db: Session,
    user_data: dict,
    chi_nhanh: str = "",
    search: str = "",
    trang_thai: str = "",
    han_hoan_thanh: str = ""
):
    """
    Hàm helper để xây dựng và trả về câu truy vấn SQLAlchemy cho các công việc
    dựa trên các bộ lọc được cung cấp. Việc xác định chi nhánh nào cần lọc
    (dựa trên GPS hay form) đã được thực hiện ở hàm `home`.
    """
    role = user_data.get("role")

    tasks_query = db.query(Task)

    # Loại bỏ công việc đã xoá cho các vai trò không phải quản lý cấp cao
    if role not in ["quanly", "ktv", "admin", "boss"]:
        tasks_query = tasks_query.filter(Task.trang_thai != "Đã xoá")

    # Lọc theo chi nhánh (nếu có).
    # `chi_nhanh` ở đây đã được xác định một cách chính xác ở hàm `home`
    # (là chi nhánh GPS cho lễ tân, hoặc chi nhánh từ bộ lọc cho các role khác).
    if chi_nhanh:
        tasks_query = tasks_query.filter(Task.chi_nhanh == chi_nhanh)

    # Lọc theo từ khóa
    if search:
        clean_search = re.sub(r'\s+', ' ', search).strip()
        search_pattern = f"%{clean_search}%"
        tasks_query = tasks_query.filter(
            or_(
                Task.chi_nhanh.ilike(search_pattern),
                Task.phong.ilike(search_pattern),
                Task.mo_ta.ilike(search_pattern),
                Task.trang_thai.ilike(search_pattern),
                Task.nguoi_tao.ilike(search_pattern),
                Task.nguoi_thuc_hien.ilike(search_pattern),
                Task.ghi_chu.ilike(search_pattern),
            )
        )

    # Lọc theo trạng thái
    if trang_thai:
        tasks_query = tasks_query.filter(Task.trang_thai == trang_thai)

    # Lọc theo hạn hoàn thành
    if han_hoan_thanh:
        try:
            han_date = datetime.strptime(han_hoan_thanh, "%Y-%m-%d").date()
            tasks_query = tasks_query.filter(func.date(Task.han_hoan_thanh) == han_date)
        except (ValueError, TypeError):
            pass

    return tasks_query

@app.get("/home", response_class=HTMLResponse)
def home(
    request: Request,
    chi_nhanh: str = "",
    search: str = "",
    trang_thai: str = "",
    han_hoan_thanh: str = "",
    page: int = 1,
    per_page: int = 8,
    db: Session = Depends(get_db)
):
    user_data = request.session.get("user")
    today = datetime.now(VN_TZ)

    if not user_data:
        return RedirectResponse("/login", status_code=303)

    # Đồng bộ trạng thái "Quá hạn" trước khi query dữ liệu
    try:
        db.query(Task).filter(
            Task.trang_thai == "Đang chờ",
            Task.han_hoan_thanh < today
        ).update({"trang_thai": "Quá hạn"}, synchronize_session=False)
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f"Lỗi khi cập nhật trạng thái công việc quá hạn trong route /home: {e}", exc_info=True)

    username = user_data["code"]
    role = user_data["role"]
    user_name = user_data["name"]

    # Lấy chi nhánh hoạt động từ GPS (active_branch) hoặc chi nhánh mặc định của user
    # --- LOGIC MỚI ĐỂ LẤY CHI NHÁNH ---
    # 1. Lấy thông tin user đầy đủ từ DB để có last_active_branch
    user_from_db = db.query(User).filter(User.code == username).first()

    # 2. Xác định chi nhánh hoạt động theo thứ tự ưu tiên:
    #    - Ưu tiên 1: Chi nhánh từ session (vừa quét GPS trong phiên này).
    #    - Ưu tiên 2: Chi nhánh hoạt động cuối cùng đã lưu trong DB.
    #    - Ưu tiên 3: Chi nhánh mặc định của user (fallback).
    active_branch = (
        request.session.get("active_branch")
        or (user_from_db.last_active_branch if user_from_db and hasattr(user_from_db, 'last_active_branch') else None)
        or user_data.get("branch")
    )

    # Xác định chi nhánh để lọc query dựa trên vai trò
    branch_to_filter = ""

    if role == 'letan':
        # ✅ Lễ tân: luôn ưu tiên GPS (active_branch) nếu có
        if request.session.get("active_branch"):
            branch_to_filter = chi_nhanh or request.session["active_branch"]
        else:
            branch_to_filter = chi_nhanh or (user_from_db.last_active_branch if user_from_db and user_from_db.last_active_branch else user_data.get("branch"))
    else:
        # ✅ Quản lý, KTV, Admin, Boss: chỉ lọc khi chọn từ form
        branch_to_filter = chi_nhanh


    # ✅ Query công việc với chi nhánh đã được xác định
    tasks_query = _get_filtered_tasks_query(
        db, user_data, branch_to_filter, search, trang_thai, han_hoan_thanh
    )

    # ✅ Lấy tất cả công việc cho Lịch (bỏ qua filter ngày và phân trang)
    # Điều này đảm bảo lịch luôn hiển thị tất cả các công việc phù hợp với bộ lọc hiện tại,
    # không bị giới hạn bởi bộ lọc ngày cụ thể.
    calendar_tasks_query = _get_filtered_tasks_query(
        db, user_data, branch_to_filter, search, trang_thai, "" # han_hoan_thanh rỗng
    )
    all_tasks_for_cal = calendar_tasks_query.all()
    calendar_tasks_data = []
    for t in all_tasks_for_cal:
        # Dữ liệu đã được đồng bộ ở đầu route, chỉ cần lấy trực tiếp
        calendar_tasks_data.append({
            "id": t.id,
            "phong": t.phong,
            "mo_ta": t.mo_ta,
            "han_hoan_thanh": format_datetime_display(t.han_hoan_thanh, with_time=False),
            "han_hoan_thanh_raw": t.han_hoan_thanh.isoformat() if t.han_hoan_thanh else None,
            "trang_thai": t.trang_thai,
        })

    # ✅ Tổng số task
    total_tasks = tasks_query.count()
    total_pages = max(1, (total_tasks + per_page - 1) // per_page)

    # ✅ Sắp xếp
    order = {"Quá hạn": 0, "Đang chờ": 1, "Hoàn thành": 2, "Đã xoá": 3}
    rows = (
        tasks_query.order_by(
            case(order, value=Task.trang_thai, else_=99),
            Task.han_hoan_thanh.nullslast(),
        )
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    # ✅ Chuẩn bị dữ liệu
    tasks, chi_nhanhs_set = [], set()
    for t in rows:
        chi_nhanhs_set.add(t.chi_nhanh)

        # Dữ liệu đã được đồng bộ ở đầu route, chỉ cần lấy trực tiếp

        tasks.append({
            "id": t.id,
            "chi_nhanh": t.chi_nhanh,
            "phong": t.phong,
            "mo_ta": t.mo_ta,
            "ngay_tao": format_datetime_display(t.ngay_tao, with_time=True),
            "han_hoan_thanh": format_datetime_display(t.han_hoan_thanh, with_time=False),
            "han_hoan_thanh_raw": t.han_hoan_thanh.isoformat() if t.han_hoan_thanh else None,
            "trang_thai": t.trang_thai,
            "nguoi_tao": t.nguoi_tao,
            "ghi_chu": t.ghi_chu or "",
            "nguoi_thuc_hien": t.nguoi_thuc_hien,
            "ngay_hoan_thanh": format_datetime_display(t.ngay_hoan_thanh, with_time=True) if t.ngay_hoan_thanh else "",
            "is_overdue": t.trang_thai == "Quá hạn",
        })

    # ✅ Thống kê (đã được đơn giản hóa vì DB đã đồng bộ)
    thong_ke = {
        "tong_cong_viec": total_tasks,
        "hoan_thanh": tasks_query.filter(Task.trang_thai == "Hoàn thành").count(),
        "hoan_thanh_tuan": tasks_query.filter(
            Task.trang_thai == "Hoàn thành",
            Task.ngay_hoan_thanh >= today.replace(hour=0, minute=0) - timedelta(days=today.weekday()),
        ).count(),
        "hoan_thanh_thang": tasks_query.filter(
            Task.trang_thai == "Hoàn thành", func.extract("month", Task.ngay_hoan_thanh) == today.month
        ).count(),
        "dang_cho": tasks_query.filter(Task.trang_thai == "Đang chờ").count(),
        "qua_han": tasks_query.filter(Task.trang_thai == "Quá hạn").count(),
    }

    if role in ["admin", "boss"]:
        # Admin/Boss: luôn thấy tất cả chi nhánh
        chi_nhanhs_display = BRANCHES
    else:
        chi_nhanhs_display = sorted(chi_nhanhs_set)


    # Tạo query string cho phân trang, giữ lại các bộ lọc hiện tại
    query_params = {
        "chi_nhanh": branch_to_filter,
        "search": search,
        "trang_thai": trang_thai,
        "han_hoan_thanh": han_hoan_thanh,
        "per_page": per_page,
    }
    # Loại bỏ các key có giá trị rỗng hoặc None
    active_filters = {k: v for k, v in query_params.items() if v}
    pagination_query_string = urlencode(active_filters)

    # ✅ Render template
    response = templates.TemplateResponse(
        "home.html",
        {
            "request": request,
            "tasks": tasks,
            "user": username,
            "role": role,
            "user_name": user_name,
            "search": search,
            "trang_thai": trang_thai,
            "chi_nhanh": branch_to_filter, # Sử dụng chi nhánh đã lọc để hiển thị trên dropdown
            "chi_nhanhs": chi_nhanhs_display,
            "user_chi_nhanh": active_branch,
            "branches": BRANCHES,
            "now": today,
            "thong_ke": thong_ke,
            "page": page,
            "total_pages": total_pages,
            "per_page": per_page,
            "query_string": f"&{pagination_query_string}" if pagination_query_string else "",
            "all_tasks_for_calendar": calendar_tasks_data,
        },
    )
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

@app.post("/add")
def add_task(
    request: Request,
    chi_nhanh: Optional[str] = Form(None),
    phong: str = Form(...),
    mo_ta: str = Form(...),
    han_hoan_thanh: str = Form(...),
    nguoi_tao: str = Form(...),
    ghi_chu: str = Form(""),
    db: Session = Depends(get_db)
):
    user = request.session.get("user")
    if not user:
        return RedirectResponse("/login", status_code=303)

    role = user.get("role")

    # Đối với các vai trò không có dropdown chọn chi nhánh (ví dụ: lễ tân),
    # chi nhánh phải được xác định một cách an toàn ở phía server
    # dựa trên chi nhánh hoạt động (active_branch) từ GPS/session.
    if role not in ["quanly", "ktv", "admin", "boss"]:
        user_from_db = db.query(User).filter(User.code == user["code"]).first()
        # Sử dụng logic tương tự như route GET /home để đảm bảo tính nhất quán
        active_branch = (
            request.session.get("active_branch")
            or (user_from_db.last_active_branch if user_from_db and hasattr(user_from_db, 'last_active_branch') else None)
            or user.get("branch")
        )
        chi_nhanh = active_branch

    if not chi_nhanh:
        raise HTTPException(status_code=400, detail="Không xác định được chi nhánh")

    han = parse_datetime_input(han_hoan_thanh)
    if han and han.tzinfo is None:
        han = VN_TZ.localize(han)
    now = datetime.now(VN_TZ)
    trang_thai = "Quá hạn" if han < now else "Đang chờ"

    new_task = Task(
        chi_nhanh=chi_nhanh,
        phong=phong,
        mo_ta=mo_ta,
        ngay_tao=now,
        han_hoan_thanh=han,
        trang_thai=trang_thai,
        nguoi_tao=nguoi_tao,
        ghi_chu=ghi_chu
    )
    db.add(new_task)
    db.commit()

    raw_query = request.scope.get("query_string", b"").decode()
    clean_query = clean_query_string(raw_query)
    redirect_url = f"/home?{clean_query}&success=1&action=add" if clean_query else "/home?success=1&action=add"

    return RedirectResponse(redirect_url, status_code=303)

@app.post("/complete/{task_id}")
async def complete_task(task_id: int, request: Request, nguoi_thuc_hien: str = Form(...), db: Session = Depends(get_db)):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Không tìm thấy công việc")

    task.trang_thai = "Hoàn thành"
    task.nguoi_thuc_hien = nguoi_thuc_hien
    task.ngay_hoan_thanh = datetime.now(VN_TZ)
    db.commit()

    if request.query_params.get("json") == "1":
        return JSONResponse({"success": True, "task_id": task_id})

    raw_query = request.scope.get("query_string", b"").decode()
    clean_query = clean_query_string(raw_query)
    redirect_url = f"/home?{clean_query}&success=1&action=complete" if clean_query else "/home?success=1&action=complete"
    return RedirectResponse(redirect_url, status_code=303)

@app.post("/delete/{task_id}")
async def delete_task(task_id: int, request: Request, db: Session = Depends(get_db)):
    user_code = request.session.get("user", {}).get("code", "")
    user = db.query(User).filter(User.code == user_code).first()
    if not user:
        return RedirectResponse("/login", status_code=303)

    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Không tìm thấy công việc")

    if user.role in ["quanly", "admin", "boss"]:
        db.delete(task)
    else:
        task.trang_thai = "Đã xoá"
    db.commit()

    if request.query_params.get("json") == "1":
        return JSONResponse({"success": True, "task_id": task_id})

    raw_query = request.scope.get("query_string", b"").decode()
    clean_query = clean_query_string(raw_query)
    redirect_url = f"/home?{clean_query}&success=1&action=delete" if clean_query else "/home?success=1&action=delete"
    return RedirectResponse(redirect_url, status_code=303)

from typing import Optional

@app.post("/edit/{task_id}")
async def edit_submit(
    request: Request,
    task_id: int,
    chi_nhanh: Optional[str] = Form(None),
    phong: str = Form(...),
    mo_ta: str = Form(...),
    han_hoan_thanh: str = Form(...),
    ghi_chu: str = Form(""),
    db: Session = Depends(get_db)
):
    # Lấy session user
    user = request.session.get("user")
    if not user:
        return RedirectResponse("/login", status_code=303)

    role = user.get("role")

    # Tương tự như khi thêm mới, vai trò không có dropdown chọn chi nhánh
    # phải được xác định chi nhánh một cách an toàn ở phía server.
    if role not in ["quanly", "ktv", "admin", "boss"]:
        user_from_db = db.query(User).filter(User.code == user["code"]).first()
        active_branch = (
            request.session.get("active_branch")
            or (user_from_db.last_active_branch if user_from_db and hasattr(user_from_db, 'last_active_branch') else None)
            or user.get("branch")
        )
        chi_nhanh = active_branch

    if not chi_nhanh:
        raise HTTPException(status_code=400, detail="Không xác định được chi nhánh")

    # Tìm công việc cần sửa
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Không tìm thấy công việc")

    # Parse hạn hoàn thành
    han = parse_datetime_input(han_hoan_thanh)
    now = datetime.now(VN_TZ)

    # Cập nhật dữ liệu
    task.chi_nhanh = chi_nhanh
    task.phong = phong
    task.mo_ta = mo_ta
    task.han_hoan_thanh = han
    task.ghi_chu = ghi_chu
    task.trang_thai = "Quá hạn" if han < now else "Đang chờ"

    db.commit()

    # Lấy query_string để redirect giữ lại filter
    form_data = await request.form()
    redirect_query = form_data.get("redirect_query", "")

    return RedirectResponse(f"/home?success=1&action=update{('&' + redirect_query) if redirect_query else ''}", status_code=303)

@app.get("/send-overdue-alerts")
async def send_overdue_alerts(request: Request, db: Session = Depends(get_db)):
    try:
        now = datetime.now(VN_TZ)

        # Cập nhật trạng thái "Quá hạn"
        overdue_to_update = db.query(Task).filter(
            Task.trang_thai == "Đang chờ",
            Task.han_hoan_thanh < now
        ).all()
        for task in overdue_to_update:
            task.trang_thai = "Quá hạn"
        if overdue_to_update:
            db.commit()

        # Lấy công việc quá hạn
        tasks = db.query(Task).filter(
            Task.trang_thai == "Quá hạn"
        ).order_by(Task.chi_nhanh.asc(), Task.phong.asc()).all()

        if not tasks:
            return JSONResponse({"message": "Không có công việc quá hạn."})

        from collections import defaultdict
        grouped = defaultdict(list)
        for t in tasks:
            grouped[t.chi_nhanh].append(t)

        base_url = str(request.base_url).rstrip("/")
        total_sent = 0

        for chi_nhanh, task_list in grouped.items():
            # Bảng HTML có kẻ dòng, căn giữa tiêu đề, chữ căn đều
            rows_html = "\n".join([
                f"""
                <tr style="border-bottom:1px solid #e5e7eb;">
                    <td style="padding:10px;">{t.phong}</td>
                    <td style="padding:10px;">
                        <a href="{base_url}/edit/{t.id}" target="_blank" style="color:#2563eb; text-decoration:none;">
                            {t.mo_ta}
                        </a>
                    </td>
                    <td style="padding:10px;">{t.ghi_chu or ''}</td>
                </tr>
                """ for t in task_list
            ])

            subject = f"🕹 CẢNH BÁO: {len(task_list)} công việc quá hạn tại {chi_nhanh}"

            body = f"""
            <html>
            <body style="font-family:Segoe UI, sans-serif; font-size:15px; color:#1f2937; background-color:#f9fafb; padding:24px;">
                <div style="max-width:700px; margin:auto; background:white; padding:24px; border-radius:8px; border:1px solid #e5e7eb; text-align:justify;">
                    <h2 style="color:#dc2626; font-weight:600; margin-bottom:16px; font-size:20px; text-align:center;">
                        {chi_nhanh} CẢNH BÁO CÔNG VIỆC QUÁ HẠN
                    </h2>

                    <p style="font-size:15px; line-height:1.6;">
                        🍀 Hệ thống ghi nhận có <strong>{len(task_list)} công việc</strong> tại chi nhánh <strong>{chi_nhanh}</strong> đang quá hạn xử lý.<br></p>
                    
                    <table style="
                        width:100%;
                        border-collapse:collapse;
                        margin-top:20px;
                        font-size:14px;
                        background-color:white;
                        box-shadow:0 0 8px rgba(0,0,0,0.04);
                    ">
                        <thead style="background-color: #f3f4f6;">
                            <tr style="border-bottom:1px solid #d1d5db;">
                                <th style="text-align:center; padding:10px; border:1px solid #d1d5db;">Phòng</th>
                                <th style="text-align:center; padding:10px; border:1px solid #d1d5db;">Mô tả</th>
                                <th style="text-align:center; padding:10px; border:1px solid #d1d5db;">Ghi chú</th>
                            </tr>
                        </thead>
                        <tbody>
                            {rows_html}
                        </tbody>
                    </table>

                    <p style="font-size:15px; line-height:1.6;"> ❗️ Vui lòng kiểm tra và xử lý kịp thời để đảm bảo tiến độ công việc. </p>

                    <p style="margin-top:16px; font-size:13px; color:#9ca3af;">
                        (Email tự động từ hệ thống quản lý công việc Bin Bin Hotel.)
                    </p>
                </div>
            </body>
            </html>
            """

            await send_alert_email(ALERT_EMAIL, subject, body, html=True)
            total_sent += len(task_list)

        return JSONResponse({"sent_total": total_sent, "branches": list(grouped.keys())})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"error": str(e)}, status_code=500)
 
from employees import employees
from database import SessionLocal
from models import User
def sync_employees_from_source(db: Session, employees: list[dict], force_delete: bool = False):
    allowed_login_roles = ["letan", "quanly", "ktv", "admin", "boss"]
    seen_codes = set()

    # --- Bước 1: Thu thập danh sách code từ employees ---
    incoming_codes = set()
    for emp in employees:
        code = emp.get("code", "").strip()
        if code:
            incoming_codes.add(code)

    # --- Bước 2: Xóa nhân viên không còn trong employees (nếu force_delete=True) ---
    if force_delete:
        db.query(User).filter(~User.code.in_(incoming_codes)).delete(synchronize_session=False)
        db.commit()
        logger.info("[SYNC] Đã xóa các nhân viên không còn trong danh sách nguồn.")

    # --- Bước 3: Đồng bộ từng nhân viên ---
    for emp in employees:
        code = emp.get("code", "").strip()
        if not code or code in seen_codes:
            continue
        seen_codes.add(code)

        name = emp.get("name", "").strip()
        branch = emp.get("branch", "").strip()
        role = emp.get("role", "").strip()

        # Nếu chưa có role thì suy từ code
        if not role:
            role_map = {"LT": "letan", "BP": "buongphong", "BV": "baove",
                        "QL": "quanly", "KTV": "ktv"}
            role = next((v for k, v in role_map.items() if k in code.upper()), "khac")
            if code.lower() in ["admin", "boss"]:
                role = code.lower()

        existing = db.query(User).filter(User.code == code).first()
        if existing:
            # Cập nhật thông tin khác (KHÔNG reset password)
            existing.name = name
            existing.role = role
            existing.branch = branch
        else:
            # Tạo user mới → set mật khẩu một lần
            password = emp.get("password")
            if not password:
                password = "999" if role in allowed_login_roles else ""
            db.add(User(code=code, name=name, password=password, role=role, branch=branch))

    db.commit() # Commit một lần duy nhất ở cuối hàm

    logger.info("Đồng bộ nhân viên thành công")

@app.get("/sync-employees")
def sync_employees_endpoint(request: Request):
    """
    Endpoint để đồng bộ lại dữ liệu nhân viên từ employees.py vào database.
    Chỉ cho phép admin hoặc boss thực hiện.
    """
    user = request.session.get("user")
    if not user or user.get("role") not in ["admin", "boss"]:
        raise HTTPException(status_code=403, detail="Chỉ admin hoặc boss mới được đồng bộ nhân viên.")
    db = SessionLocal()
    sync_employees_from_source(db=db, employees=employees, force_delete=True)
    db.close()
    return {"status": "success", "message": "Đã đồng bộ lại danh sách nhân viên từ employees.py"}

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=303)

from sqlalchemy import text
from database import SessionLocal, init_db
import os, time, threading, atexit
from apscheduler.schedulers.background import BackgroundScheduler
from datetime import datetime
from employees import employees

def update_overdue_tasks_status():
    """
    Tác vụ nền tự động cập nhật trạng thái các công việc từ "Đang chờ" sang "Quá hạn".
    Sử dụng một câu lệnh UPDATE duy nhất để tối ưu hiệu suất và bộ nhớ.
    So sánh thời gian đầy đủ (datetime) để đảm bảo tính chính xác.
    """
    db = SessionLocal()
    try:
        # Lấy thời gian hiện tại theo múi giờ Việt Nam
        now_vn = datetime.now(VN_TZ)
        
        # Thực hiện một câu lệnh UPDATE trực tiếp trên DB.
        # So sánh thời gian đầy đủ của han_hoan_thanh với thời gian hiện tại.
        # Một công việc được coi là quá hạn nếu thời gian hiện tại đã vượt qua hạn hoàn thành.
        updated_count = db.query(Task).filter(
            Task.trang_thai == "Đang chờ",
            Task.han_hoan_thanh < now_vn
        ).update({"trang_thai": "Quá hạn"}, synchronize_session=False)
        
        db.commit()

        if updated_count > 0:
            logger.info(f"[AUTO_UPDATE_STATUS] Đã cập nhật {updated_count} công việc sang trạng thái 'Quá hạn'.")
        else:
            logger.info("[AUTO_UPDATE_STATUS] Không có công việc nào cần cập nhật trạng thái.")

    except Exception as e:
        logger.error(f"[AUTO_UPDATE_STATUS] Lỗi khi cập nhật trạng thái công việc quá hạn: {e}", exc_info=True)
        db.rollback()
    finally:
        db.close()

from database import SessionLocal, init_db
import os, time, threading, atexit
from apscheduler.schedulers.background import BackgroundScheduler
from datetime import datetime
from employees import employees

# Danh sách các bảng có cột id SERIAL cần reset sequence
TABLES_WITH_SERIAL_ID = ["tasks", "attendance_log", "attendance_records", "service_records", "lost_and_found_items"]

def reset_sequence(db, table_name: str, id_col: str = "id"):
    """
    Reset sequence cho bảng cụ thể, đảm bảo id không bị trùng.
    """
    seq_name = f"{table_name}_{id_col}_seq"
    sql = f"SELECT setval('{seq_name}', (SELECT COALESCE(MAX({id_col}), 0) + 1 FROM {table_name}), false)"
    try:
        db.execute(text(sql))
        db.commit()
        logger.info(f"Đã đồng bộ sequence cho bảng {table_name}")
    except Exception as e:
        logger.error(f"Lỗi khi reset sequence cho {table_name}: {e}", exc_info=True)

@app.on_event("startup")
def startup():
    logger.info("🚀 Khởi động ứng dụng...")

    # --- 1. Init DB ---
    init_db()

    # --- 2. Reset sequence cho các bảng ---
    with SessionLocal() as db:
        for table in TABLES_WITH_SERIAL_ID:
            reset_sequence(db, table)

        # --- 3. Đồng bộ nhân viên (chạy 1 lần khi startup) ---
        try:
            sync_employees_from_source(db=db, employees=employees, force_delete=False)
            logger.info("Hoàn tất đồng bộ nhân viên từ employees.py")
        except Exception as e:
            logger.error("Không thể đồng bộ nhân viên", exc_info=True)

    # --- 4. Lập lịch các tác vụ nền ---
    def auto_logout_job():
        logger.info("Kích hoạt đăng xuất tự động cho tất cả client.")

    scheduler = BackgroundScheduler(timezone=str(VN_TZ))
    # Chạy tác vụ cập nhật trạng thái công việc mỗi 3 giờ để tiết kiệm tài nguyên.
    # Điều này đảm bảo hệ thống phản ứng nhanh hơn và linh hoạt hơn với các
    # nền tảng hosting có thể "ngủ" (sleep) khi không có traffic.
    # misfire_grace_time=600: Nếu job bị lỡ, nó vẫn sẽ chạy nếu server thức dậy trong vòng 10 phút.
    # Vô hiệu hóa tác vụ nền cập nhật trạng thái vì đã có cập nhật tức thì trong route /home.
    # scheduler.add_job(update_overdue_tasks_status, 'interval', hours=3, id='update_overdue_tasks', misfire_grace_time=600)
    scheduler.add_job(auto_logout_job, 'cron', hour=6, minute=59)
    scheduler.add_job(auto_logout_job, 'cron', hour=18, minute=59)
    scheduler.add_job(run_daily_absence_check, 'cron', hour=7, minute=0, misfire_grace_time=900)
    scheduler.start()

    # --- 5. Shutdown scheduler khi app stop ---
    atexit.register(lambda: scheduler.shutdown())

    logger.info("✅ Startup hoàn tất: DB init, reset sequence, sync nhân viên, lập lịch các tác vụ nền.")

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run("main:app", host="0.0.0.0", port=port)

def get_lan_ip():
    """Hàm này lấy địa chỉ IP nội bộ (LAN) của máy chủ."""
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        # Không cần phải kết nối được, chỉ là một mẹo để lấy IP
        s.connect(('10.255.255.255', 1))
        IP = s.getsockname()[0]
    except Exception:
        IP = '127.0.0.1' # Fallback nếu không lấy được IP
    finally:
        s.close()
    return IP

def get_active_branch(request: Request, db: Session, user_data: dict) -> Optional[str]:
    """
    Xác định chi nhánh hoạt động của người dùng theo thứ tự ưu tiên:
    1. Chi nhánh từ session (vừa quét GPS trong phiên này).
    2. Chi nhánh hoạt động cuối cùng đã lưu trong DB.
    3. Chi nhánh mặc định của user (fallback).
    """
    username = user_data.get("code")
    if not username:
        return None

    # 1. Lấy từ session (ưu tiên cao nhất)
    active_branch = request.session.get("active_branch")
    if active_branch:
        return active_branch

    # 2. Lấy từ DB
    user_from_db = db.query(User).filter(User.code == username).first()
    if user_from_db and hasattr(user_from_db, 'last_active_branch') and user_from_db.last_active_branch:
        return user_from_db.last_active_branch

    # 3. Lấy từ chi nhánh mặc định
    return user_data.get("branch")

@app.get("/api/lost-and-found", response_class=JSONResponse)
async def api_lost_and_found_items(
    request: Request,
    db: Session = Depends(get_db),
    page: int = 1,
    per_page: int = 20,
    search: Optional[str] = None,
    status: Optional[str] = None,
    chi_nhanh: Optional[str] = None,
    sort_by: Optional[str] = 'found_date',
    sort_order: Optional[str] = 'desc',
    found_date: Optional[str] = None,
    reported_by: Optional[str] = None,
):
    user_data = request.session.get("user")
    if not user_data: # Mở quyền cho tất cả user đã đăng nhập
        raise HTTPException(status_code=403, detail="Unauthorized")

    query = db.query(LostAndFoundItem)

    role = user_data.get("role")
    # Tất cả các vai trò không phải admin/boss sẽ không thấy các mục đã bị "xóa mềm"
    if role not in ["admin", "boss"]:
        query = query.filter(LostAndFoundItem.status != LostItemStatus.DELETED)

    role = user_data.get("role")
    branch_to_filter = ""
    if role == 'letan':
        active_branch = get_active_branch(request, db, user_data)
        branch_to_filter = chi_nhanh or active_branch
    else: # quanly, admin, boss
        branch_to_filter = chi_nhanh

    if branch_to_filter:
        query = query.filter(LostAndFoundItem.chi_nhanh == branch_to_filter)

    if search:
        search_pattern = f"%{search}%"
        query = query.filter(or_(
            LostAndFoundItem.item_name.ilike(search_pattern),
            LostAndFoundItem.description.ilike(search_pattern),
            LostAndFoundItem.found_location.ilike(search_pattern),
            LostAndFoundItem.recorded_by.ilike(search_pattern),
            LostAndFoundItem.owner_name.ilike(search_pattern),
            LostAndFoundItem.reported_by.ilike(search_pattern),
        ))

    if status:
        try:
            # Chuyển đổi chuỗi status thành enum member để đảm bảo type safety
            status_enum = LostItemStatus(status)
            query = query.filter(LostAndFoundItem.status == status_enum)
        except ValueError:
            pass # Bỏ qua nếu status không hợp lệ

    if found_date:
        try:
            # 1. Parse ngày người dùng chọn từ bộ lọc.
            parsed_date = datetime.strptime(found_date, "%Y-%m-%d").date()

            # 2. Tạo khoảng thời gian bắt đầu và kết thúc của ngày đó THEO MÚI GIỜ VIỆT NAM.
            # Bắt đầu từ 00:00:00 của ngày đã chọn.
            start_of_day_vn = VN_TZ.localize(datetime.combine(parsed_date, datetime.min.time()))
            # Kết thúc vào 23:59:59 của ngày đã chọn.
            end_of_day_vn = VN_TZ.localize(datetime.combine(parsed_date, datetime.max.time()))

            # 3. Lọc các bản ghi có found_date nằm trong khoảng thời gian trên.
            # SQLAlchemy sẽ tự động xử lý việc chuyển đổi múi giờ khi so sánh với DB.
            query = query.filter(LostAndFoundItem.found_date.between(start_of_day_vn, end_of_day_vn))
        except (ValueError, TypeError):
            pass # Bỏ qua nếu định dạng ngày không hợp lệ

    if reported_by:
        query = query.filter(LostAndFoundItem.reported_by.ilike(f"%{reported_by}%"))

    # Sắp xếp theo logic mới:
    # 1. Ưu tiên trạng thái: Đang lưu giữ > Đã trả khách > Thanh lý > Đã xoá
    # 2. Sau đó theo ngày mới nhất (ưu tiên ngày trả, fallback về ngày tìm thấy)
    status_order = case(
        {
            LostItemStatus.STORED: 1,
            LostItemStatus.RETURNED: 2,
            LostItemStatus.DISPOSED: 3,
            LostItemStatus.DELETED: 4,
        },
        value=LostAndFoundItem.status,
        else_=99
    )
    sort_expression = desc(func.coalesce(LostAndFoundItem.return_date, LostAndFoundItem.found_date))

    # Apply sorting
    query = query.order_by(asc(status_order), sort_expression)

    total_records = query.count()
    items = query.offset((page - 1) * per_page).limit(per_page).all()

    return {"records": jsonable_encoder(items), "currentPage": page, "totalPages": math.ceil(total_records / per_page), "totalRecords": total_records}

class BatchDeleteLostItemsPayload(BaseModel):
    ids: List[int]


# --- Lost & Found Endpoints ---

@app.get("/lost-and-found", response_class=HTMLResponse)
async def lost_and_found_page(
    request: Request,
    db: Session = Depends(get_db),
    search: Optional[str] = None,
    status: Optional[str] = None,
    chi_nhanh: Optional[str] = None
):
    user_data = request.session.get("user")
    if not user_data: # Mở quyền cho tất cả user đã đăng nhập
        return RedirectResponse("/choose-function", status_code=303)

    role = user_data.get("role")
    if role == 'letan':
        active_branch = get_active_branch(request, db, user_data)
    else:
        active_branch = ""
    
    return templates.TemplateResponse("lost_and_found.html", {
        "request": request,
        "user": user_data,
        "statuses": [s.value for s in LostItemStatus if s != LostItemStatus.DELETED],
        "initial_branch_filter": active_branch,
        "branches": BRANCHES,
        "branch_filter": chi_nhanh,
        "status_filter": status,
    })

@app.post("/lost-and-found/add", response_class=JSONResponse)
async def add_lost_item(request: Request, db: Session = Depends(get_db)):
    user_data = request.session.get("user")
    if not user_data:
        raise HTTPException(status_code=403, detail="Unauthorized")

    form_data = await request.form()
    role = user_data.get("role")
    chi_nhanh_form = form_data.get("chi_nhanh")

    # Lễ tân sẽ tự động gán chi nhánh hiện tại. Các role khác có thể chọn.
    if role == 'letan':
        active_branch = get_active_branch(request, db, user_data)
        item_branch = active_branch
    else:
        item_branch = chi_nhanh_form

    # Admin/Boss có thể ghi nhận cho người khác, các role khác tự động ghi nhận là chính mình.
    recorded_by_code = user_data.get("code") # Mặc định là người đang đăng nhập
    if role in ['admin', 'boss']:
        # Lấy mã người ghi nhận từ form, nếu có
        recorded_by_from_form = form_data.get("recorded_by")
        # Chỉ ghi đè nếu có giá trị được gửi lên và không rỗng
        if recorded_by_from_form:
            recorded_by_code = recorded_by_from_form

    new_item = LostAndFoundItem(
        item_name=form_data.get("item_name"),
        description=form_data.get("description"),
        found_location=form_data.get("found_location"),
        reported_by=form_data.get("reported_by"),
        owner_name=form_data.get("owner_name"),
        owner_contact=form_data.get("owner_contact"),
        notes=form_data.get("notes"),
        chi_nhanh=item_branch,
        found_date=datetime.now(VN_TZ),
        recorded_by=recorded_by_code
    )
    db.add(new_item)
    db.commit()
    return JSONResponse({"status": "success", "message": "Đã thêm món đồ thành công."})

@app.post("/lost-and-found/update/{item_id}", response_class=JSONResponse)
async def update_lost_item(item_id: int, request: Request, db: Session = Depends(get_db)):
    user_data = request.session.get("user")
    if not user_data: # Mở quyền cho tất cả user đã đăng nhập
        raise HTTPException(status_code=403, detail="Unauthorized")

    item = db.query(LostAndFoundItem).filter(LostAndFoundItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    form_data = await request.form()
    action = form_data.get("action")

    if action == "return":
        item.status = LostItemStatus.RETURNED
        item.owner_name = form_data.get("owner_name")
        item.owner_contact = form_data.get("owner_contact")
        item.return_date = datetime.now(VN_TZ)
    elif action == "dispose":
        item.status = LostItemStatus.DISPOSED
        item.disposed_by = form_data.get("disposed_by")
        item.return_date = datetime.now(VN_TZ)
        try:
            disposed_amount_str = form_data.get("disposed_amount")
            if disposed_amount_str:
                item.disposed_amount = float(disposed_amount_str)
            else:
                item.disposed_amount = None
        except (ValueError, TypeError):
            item.disposed_amount = None

    # Chỉ cập nhật ghi chú nếu người dùng nhập nội dung mới.
    # Nếu để trống, ghi chú cũ sẽ được giữ lại để tránh mất dữ liệu.
    updated_notes = form_data.get("notes")
    if updated_notes:
        item.notes = updated_notes
    db.commit()
    return JSONResponse({"status": "success", "message": "Đã cập nhật trạng thái."})

@app.post("/lost-and-found/delete/{item_id}", response_class=JSONResponse)
async def delete_lost_item(item_id: int, request: Request, db: Session = Depends(get_db)):
    user_data = request.session.get("user")
    if not user_data: # Mở quyền cho tất cả user đã đăng nhập
        raise HTTPException(status_code=403, detail="Unauthorized")

    user_role = user_data.get("role")
    if user_role in ["admin", "boss"]:
        # Admin/Boss: Xóa vĩnh viễn (hard delete)
        item = db.query(LostAndFoundItem).filter(LostAndFoundItem.id == item_id).first()
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")
        db.delete(item)
        db.commit()
        return JSONResponse({"status": "success", "message": "Đã xóa vĩnh viễn món đồ."})
    else:
        # Các vai trò khác: Cập nhật trạng thái thành "Đã xoá" (soft delete)
        # Sử dụng query.update() để đảm bảo chỉ có một câu lệnh UPDATE được thực thi,
        # tránh các vấn đề tiềm ẩn về trạng thái của object.
        deleter_info = f"{user_data.get('name', '')} ({user_data.get('code', '')})"
        updated_count = db.query(LostAndFoundItem).filter(LostAndFoundItem.id == item_id).update(
            {
                "status": LostItemStatus.DELETED.value,
                "deleted_by": deleter_info,
                "deleted_date": datetime.now(VN_TZ)
            }, synchronize_session=False
        )
        if updated_count == 0:
            db.rollback()
            raise HTTPException(status_code=404, detail="Item not found to delete.")
        db.commit()
        return JSONResponse({"status": "success", "message": "Đã xóa món đồ thành công."})

@app.post("/lost-and-found/batch-delete", response_class=JSONResponse)
async def batch_delete_lost_items(
    payload: BatchDeleteLostItemsPayload,
    request: Request,
    db: Session = Depends(get_db)
):
    user_data = request.session.get("user")
    user_role = user_data.get("role") if user_data else None
    if not user_role: # Mở quyền cho tất cả user đã đăng nhập
        raise HTTPException(status_code=403, detail="Unauthorized")

    if not payload.ids:
        return JSONResponse({"status": "noop", "message": "Không có mục nào được chọn để xóa."})

    try:
        if user_role in ["admin", "boss"]:
            # Admin/Boss: Xóa vĩnh viễn
            num_deleted = db.query(LostAndFoundItem).filter(LostAndFoundItem.id.in_(payload.ids)).delete(synchronize_session=False)
            db.commit()
            return JSONResponse({"status": "success", "message": f"Đã xóa vĩnh viễn {num_deleted} mục."})
        else:
            # Các vai trò khác: Ẩn đi
            num_updated = db.query(LostAndFoundItem).filter(LostAndFoundItem.id.in_(payload.ids)).update({"status": LostItemStatus.DELETED.value}, synchronize_session=False)
            db.commit()
            return JSONResponse({"status": "success", "message": f"Đã xóa thành công {num_updated} mục."})
    except Exception as e:
        db.rollback()
        logger.error(f"Lỗi khi xóa hàng loạt đồ thất lạc: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Lỗi server khi xóa.")

@app.get("/api/users/search-login-users", response_class=JSONResponse)
def search_login_users(q: str = "", db: Session = Depends(get_db)):
    """API để tìm kiếm người dùng có quyền đăng nhập (lễ tân, ql, ktv, admin, boss)."""
    if not q:
        return JSONResponse(content=[])
    
    search_pattern = f"%{q}%"
    allowed_roles = ["letan", "quanly", "ktv", "admin", "boss"]
    
    users = db.query(User).filter(
        User.role.in_(allowed_roles),
        or_(
            User.code.ilike(search_pattern),
            User.name.ilike(search_pattern)
        )
    ).limit(20).all()
    
    user_list = [
        {"code": user.code, "name": user.name}
        for user in users
    ]
    return JSONResponse(content=user_list)

def get_lan_ip():
    """Hàm này lấy địa chỉ IP nội bộ (LAN) của máy chủ."""
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        # Không cần phải kết nối được, chỉ là một mẹo để lấy IP
        s.connect(('10.255.255.255', 1))
        IP = s.getsockname()[0]
    except Exception:
        IP = '127.0.0.1' # Fallback nếu không lấy được IP
    finally:
        s.close()
    return IP

@app.get("/show_qr", response_class=HTMLResponse)
async def show_qr(request: Request, db: Session = Depends(get_db)):
    user = request.session.get("pending_user") or request.session.get("user")
    if not user:
        return RedirectResponse("/login", status_code=303)

    work_date, shift = get_current_work_shift()
    # Sửa lỗi: Áp dụng logic tương tự login_submit để xác định ca làm việc
    # Đối với KTV/Quản lý, ca luôn là NULL để đảm bảo mỗi ngày chỉ có 1 record.
    shift_value = _get_log_shift_for_user(user.get("role"), shift)
    log = db.query(AttendanceLog).filter(
        AttendanceLog.user_code == user["code"], # user là dict, nên dùng user["code"]
        AttendanceLog.date == work_date,
        AttendanceLog.shift == shift_value
    ).first()

    if log:
        if log.checked_in:
            # Nếu đã check-in thì không cần show_qr nữa → đi thẳng trang chọn chức năng
            request.session["user"] = user
            request.session.pop("pending_user", None)
            return RedirectResponse("/choose-function", status_code=303) # Thêm redirect ở đây
        else:
            qr_token = log.token
    else:
        # Trường hợp này không nên xảy ra nếu luồng đăng nhập đúng, nhưng là fallback
        import uuid
        qr_token = str(uuid.uuid4())
        log = AttendanceLog(user_code=user["code"], date=work_date, shift=shift_value, token=qr_token, checked_in=False)
        db.add(log)
        db.commit()

    request.session["qr_token"] = qr_token
    
    # Lấy host và port từ request
    request_host = request.url.hostname
    port = request.url.port
    scheme = request.url.scheme

    # Nếu host là localhost, thay thế bằng IP LAN để điện thoại có thể truy cập
    if request_host in ["localhost", "127.0.0.1"]:
        lan_ip = get_lan_ip()
        base_url = f"{scheme}://{lan_ip}:{port}"
    else:
        base_url = str(request.base_url).strip("/")
    return templates.TemplateResponse("show_qr.html", {
        "request": request,
        "qr_token": qr_token,
        "base_url": base_url,
        "user": user
    })

from services.attendance_service import push_bulk_checkin

@app.post("/attendance/checkin_bulk")
async def attendance_checkin_bulk(
    request: Request,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db)
):
    validate_csrf(request)

    # Xác định xem đây là luồng điểm danh khi đăng nhập hay điểm danh thông thường
    is_login_flow = "pending_user" in request.session and "user" not in request.session

    user = request.session.get("user") or request.session.get("pending_user")
    if not user:
        raise HTTPException(status_code=403, detail="Không có quyền điểm danh.")

    raw_data = await request.json()
    if not isinstance(raw_data, list):
        raise HTTPException(status_code=400, detail="Payload phải là danh sách")

    # Lấy chi nhánh làm việc từ payload để cập nhật trạng thái.
    # Giả định tất cả record trong 1 lần gửi đều thuộc cùng 1 chi nhánh làm việc.
    active_branch_from_payload = None
    if raw_data: # Đảm bảo raw_data không rỗng
        active_branch_from_payload = raw_data[0].get("chi_nhanh_lam")

    nguoi_diem_danh_code = user.get("code")
    user_role = user.get("role")
    user_branch = user.get("branch")
    special_roles = ["quanly", "ktv", "admin", "boss"]

    # Đối với các vai trò đặc biệt (QL, KTV, admin, boss), họ chỉ điểm danh cho chính mình.
    # Chi nhánh làm việc sẽ được tự động gán bằng chi nhánh chính của họ, không cần chọn từ UI.
    if user_role in special_roles:
        active_branch_from_payload = user_branch

    normalized_data = []
    attendance_db_records = []
    service_db_records = []
    now_vn = datetime.now(VN_TZ)

    for rec in raw_data:
        # Luôn sử dụng thời gian từ server để đảm bảo tính chính xác và tránh sai lệch múi giờ từ client.
        thoi_gian_dt = now_vn
        thoi_gian_str = thoi_gian_dt.strftime("%Y-%m-%d %H:%M:%S")

        # Dữ liệu cho Google Sheets (giữ nguyên)
        normalized_data.append({
            "sheet": rec.get("sheet"),
            "thoi_gian": thoi_gian_str, # Sử dụng thời gian server đã định dạng
            "nguoi_diem_danh": nguoi_diem_danh_code,
            "ma_nv": rec.get("ma_nv"),
            "ten_nv": rec.get("ten_nv"),
            "chi_nhanh_chinh": rec.get("chi_nhanh_chinh"),
            "chi_nhanh_lam": active_branch_from_payload,
            "la_tang_ca": "x" if rec.get("la_tang_ca") else "",
            "so_cong_nv": rec.get("so_cong_nv") or 1,
            "ghi_chu": rec.get("ghi_chu", ""),
            "dich_vu": rec.get("dich_vu") or rec.get("service") or "",
            "so_phong": rec.get("so_phong") or rec.get("room_count") or "",
            "so_luong": rec.get("so_luong") or rec.get("item_count") or ""
        })

        # Phân loại record để lưu vào DB
        is_service_record = any(rec.get(key) for key in ["dich_vu", "service", "so_phong", "room_count"])

        if is_service_record:
            # Tạo bản ghi dịch vụ
            service_db_records.append(ServiceRecord(
                ngay_cham=thoi_gian_dt.date(),
                gio_cham=thoi_gian_dt.time(), # Sử dụng time() từ datetime object đã nhận đúng múi giờ
                nguoi_cham=nguoi_diem_danh_code,
                ma_nv=rec.get("ma_nv"),
                ten_nv=rec.get("ten_nv"),
                chi_nhanh_chinh=rec.get("chi_nhanh_chinh"),
                chi_nhanh_lam=active_branch_from_payload,
                la_tang_ca=bool(rec.get("la_tang_ca")),
                dich_vu=rec.get("dich_vu") or rec.get("service") or "N/A",
                so_phong=rec.get("so_phong") or rec.get("room_count") or "",
                so_luong=rec.get("so_luong") or rec.get("item_count") or "",
                ghi_chu=rec.get("ghi_chu", "")
            ))
        else:
            # Tạo bản ghi điểm danh
            attendance_db_records.append(AttendanceRecord(
                ngay_diem_danh=thoi_gian_dt.date(),
                gio_diem_danh=thoi_gian_dt.time(), # Sử dụng time() từ datetime object đã nhận đúng múi giờ
                nguoi_diem_danh=nguoi_diem_danh_code,
                ma_nv=rec.get("ma_nv"),
                ten_nv=rec.get("ten_nv"),
                chi_nhanh_chinh=rec.get("chi_nhanh_chinh"),
                chi_nhanh_lam=active_branch_from_payload,
                la_tang_ca=bool(rec.get("la_tang_ca")),
                so_cong_nv=float(rec.get("so_cong_nv") or 1.0),
                ghi_chu=rec.get("ghi_chu", "")
            ))

    # Lấy danh sách mã nhân viên BP vừa được điểm danh
    bp_codes = [
        rec.get("ma_nv") for rec in raw_data
        if "BP" in rec.get("ma_nv", "").upper()
    ]

    # Cập nhật DB cho lễ tân đang đăng nhập và lưu các bản ghi
    try:
        # Lưu các bản ghi mới vào DB
        if attendance_db_records:
            db.add_all(attendance_db_records)
        if service_db_records:
            db.add_all(service_db_records)

        # Cập nhật thông tin cho người điểm danh
        if nguoi_diem_danh_code:
            checker_user = db.query(User).filter(User.code == nguoi_diem_danh_code).first()
            if checker_user:
                checker_user.last_checked_in_bp = bp_codes
                if active_branch_from_payload and hasattr(checker_user, 'last_active_branch'):
                    checker_user.last_active_branch = active_branch_from_payload
                    request.session["active_branch"] = active_branch_from_payload

        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        logger.error(f"Lỗi khi lưu điểm danh/dịch vụ: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Lỗi khi lưu kết quả vào cơ sở dữ liệu.")

    # Chạy push_bulk_checkin ở background để ghi vào Google Sheets
    background_tasks.add_task(push_bulk_checkin, normalized_data)

    logger.info(f"{nguoi_diem_danh_code} gửi {len(normalized_data)} record (ghi DB & queue ghi Sheets)")

    # Nếu đây là lần điểm danh ngay sau khi đăng nhập (trên mobile),
    # thì hoàn tất phiên đăng nhập và trả về URL để chuyển hướng.
    if is_login_flow and nguoi_diem_danh_code:
        user_in_db = db.query(User).filter(User.code == nguoi_diem_danh_code).first()
        if user_in_db:
            # Đánh dấu bản ghi log điểm danh là đã check-in thành công
            work_date, shift = get_current_work_shift()
            shift_value = None if user_in_db.role in ["ktv", "quanly"] else shift
            log = db.query(AttendanceLog).filter_by(
                user_code=user_in_db.code,
                date=work_date,
                shift=shift_value
            ).first()
            if log:
                log.checked_in = True
                db.commit()

            # Chuyển từ pending_user sang user chính thức trong session
            request.session["user"] = {
                "code": user_in_db.code, "role": user_in_db.role,
                "branch": user_in_db.branch, "name": user_in_db.name
            }
            request.session["after_checkin"] = "choose_function"
            request.session.pop("pending_user", None)
            return {"status": "queued", "inserted": len(normalized_data), "redirect_to": "/choose-function"}

    return {"status": "queued", "inserted": len(normalized_data)}

@app.get("/api/attendance/last-checked-in-bp", response_class=JSONResponse)
def get_last_checked_in_bp(request: Request, db: Session = Depends(get_db)):
    """
    API trả về danh sách nhân viên buồng phòng mà lễ tân đã điểm danh lần cuối.
    Dùng cho nút "Tải lại" trên trang Chấm dịch vụ.
    """
    user_data = request.session.get("user")
    if not user_data:
        raise HTTPException(status_code=403, detail="Không có quyền truy cập.")

    checker_user = db.query(User).filter(User.code == user_data["code"]).first()
    if not checker_user or not checker_user.last_checked_in_bp:
        return JSONResponse(content=[])

    service_checkin_codes = checker_user.last_checked_in_bp
    
    if not service_checkin_codes:
        return JSONResponse(content=[])

    employees_from_db = db.query(User).filter(User.code.in_(service_checkin_codes)).all()
    # Sắp xếp lại theo đúng thứ tự đã điểm danh
    emp_map = {emp.code: emp for emp in employees_from_db}
    sorted_employees = [emp_map[code] for code in service_checkin_codes if code in emp_map]
    
    employee_list = [
        { "code": emp.code, "name": emp.name, "branch": emp.branch, "so_phong": "", "so_luong": "", "dich_vu": "", "ghi_chu": "" }
        for emp in sorted_employees
    ]
    return JSONResponse(content=employee_list)

@app.get("/api/attendance/results-by-checker")
async def api_get_attendance_results(
    request: Request,
    db: Session = Depends(get_db),
    page: int = 1,
    per_page: int = 20,
    filter_type: Optional[str] = None,
    filter_date: Optional[str] = None,
    filter_nhan_vien: Optional[str] = None,
    filter_chuc_vu: Optional[str] = None,
    filter_cn_lam: Optional[str] = None,
    filter_so_cong: Optional[float] = None,
    filter_tang_ca: Optional[str] = None,
    filter_ghi_chu: Optional[str] = None,
    filter_nguoi_thuc_hien: Optional[str] = None,
    filter_dich_vu: Optional[str] = None,
    filter_so_phong: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = 'desc',
):
    """
    API trả về kết quả điểm danh.
    - Đối với admin/boss: trả về tất cả kết quả.
    - Đối với các role khác: trả về kết quả do chính người dùng đó thực hiện.
    - Hỗ trợ phân trang và lọc phía server.
    """
    user = request.session.get("user")
    allowed_roles = ['letan', 'quanly', 'ktv', 'admin', 'boss']
    if not user or user.get("role") not in allowed_roles:
        raise HTTPException(status_code=403, detail="Bạn không có quyền truy cập chức năng này.")

    checker_code = user.get("code")
    user_role = user.get("role")
    if not checker_code:
        raise HTTPException(status_code=403, detail="Không tìm thấy mã người dùng.")

    EmployeeUser = aliased(User, name="employee_user")
    CheckerUser = aliased(User, name="checker_user")

    # Base query for AttendanceRecord
    att_q = select(
        AttendanceRecord.id,
        literal_column("'Điểm danh'").label("type"),
        AttendanceRecord.ngay_diem_danh.label("date_col"),
        AttendanceRecord.gio_diem_danh.label("time_col"),
        AttendanceRecord.nguoi_diem_danh.label("nguoi_thuc_hien"),
        CheckerUser.name.label("ten_nguoi_thuc_hien"),
        AttendanceRecord.ma_nv,
        AttendanceRecord.ten_nv,
        EmployeeUser.role.label("chuc_vu_raw"),
        AttendanceRecord.chi_nhanh_lam,
        AttendanceRecord.chi_nhanh_chinh,
        EmployeeUser.branch.label("employee_branch"),
        AttendanceRecord.so_cong_nv.label("so_cong"),
        AttendanceRecord.la_tang_ca,
        AttendanceRecord.ghi_chu,
        literal_column("''").label("dich_vu"),
        literal_column("''").label("so_phong"),
        literal_column("''").label("so_luong")
    ).join(
        EmployeeUser, EmployeeUser.code == AttendanceRecord.ma_nv, isouter=True
    ).join(
        CheckerUser, CheckerUser.code == AttendanceRecord.nguoi_diem_danh, isouter=True
    )

    # Base query for ServiceRecord
    svc_q = select(
        ServiceRecord.id,
        literal_column("'Dịch vụ'").label("type"),
        ServiceRecord.ngay_cham.label("date_col"),
        ServiceRecord.gio_cham.label("time_col"),
        ServiceRecord.nguoi_cham.label("nguoi_thuc_hien"),
        CheckerUser.name.label("ten_nguoi_thuc_hien"),
        ServiceRecord.ma_nv,
        ServiceRecord.ten_nv,
        EmployeeUser.role.label("chuc_vu_raw"),
        ServiceRecord.chi_nhanh_lam,
        ServiceRecord.chi_nhanh_chinh,
        EmployeeUser.branch.label("employee_branch"),
        literal_column("NULL").cast(Float).label("so_cong"),
        ServiceRecord.la_tang_ca,
        ServiceRecord.ghi_chu,
        ServiceRecord.dich_vu,
        ServiceRecord.so_phong,
        ServiceRecord.so_luong
    ).join(
        EmployeeUser, EmployeeUser.code == ServiceRecord.ma_nv, isouter=True
    ).join(
        CheckerUser, CheckerUser.code == ServiceRecord.nguoi_cham, isouter=True
    )

    # Role-based filtering
    if user_role not in ["admin", "boss"]:
        # Non-admin roles (letan, ktv, quanly) see records they created OR records about them.
        att_q = att_q.where(or_(AttendanceRecord.nguoi_diem_danh == checker_code, AttendanceRecord.ma_nv == checker_code))
        svc_q = svc_q.where(or_(ServiceRecord.nguoi_cham == checker_code, ServiceRecord.ma_nv == checker_code))

    # Union the two queries
    u = union_all(att_q, svc_q).alias("u")

    # Build final query from the union
    final_query = select(u)

    # Apply filters
    if filter_type:
        final_query = final_query.where(u.c.type == filter_type)
    if filter_date:
        try:
            parsed_date = datetime.strptime(filter_date, "%Y-%m-%d").date()
            final_query = final_query.where(u.c.date_col == parsed_date)
        except ValueError:
            pass # Ignore invalid date format
    if filter_nhan_vien:
        search_pattern = f"%{filter_nhan_vien}%"
        final_query = final_query.where(or_(
            u.c.ma_nv.ilike(search_pattern),
            u.c.ten_nv.ilike(search_pattern)
        ))
    if filter_chuc_vu:
        matching_roles = [
            role for role, vn_role in ROLE_MAP.items()
            if filter_chuc_vu.lower() in vn_role.lower()
        ]
        if matching_roles:
            final_query = final_query.where(u.c.chuc_vu_raw.in_(matching_roles))
    if filter_cn_lam:
        final_query = final_query.where(u.c.chi_nhanh_lam == filter_cn_lam)
    if filter_so_cong is not None:
        final_query = final_query.where(u.c.so_cong == filter_so_cong)
    if filter_tang_ca and filter_tang_ca != 'all':
        is_overtime = filter_tang_ca == 'yes'
        final_query = final_query.where(u.c.la_tang_ca == is_overtime)
    if filter_ghi_chu:
        final_query = final_query.where(u.c.ghi_chu.ilike(f"%{filter_ghi_chu}%"))
    if filter_nguoi_thuc_hien and user_role in ['admin', 'boss']:
        final_query = final_query.where(or_(
            u.c.nguoi_thuc_hien.ilike(f"%{filter_nguoi_thuc_hien}%"),
            u.c.ten_nguoi_thuc_hien.ilike(f"%{filter_nguoi_thuc_hien}%")
        ))
    if filter_dich_vu:
        final_query = final_query.where(u.c.dich_vu.ilike(f"%{filter_dich_vu}%"))
    if filter_so_phong:
        final_query = final_query.where(u.c.so_phong.ilike(f"%{filter_so_phong}%"))

    # Get total count for pagination
    count_query = select(func.count()).select_from(final_query.alias("count_alias"))
    total_records = db.execute(count_query).scalar_one() or 0

    total_pages = math.ceil(total_records / per_page) if per_page > 0 else 1

    # Apply sorting
    order_expressions = []

    # Custom sort for 'chi_nhanh_lam' to follow the order in BRANCHES list
    branch_order_whens = {branch: i for i, branch in enumerate(BRANCHES)}
    chi_nhanh_lam_sort_expr = case(branch_order_whens, value=u.c.chi_nhanh_lam, else_=len(BRANCHES))

    sort_map = {
        "thoi_gian": [u.c.date_col, u.c.time_col],
        "nguoi_thuc_hien": [u.c.ten_nguoi_thuc_hien],
        "ma_nv": [u.c.ma_nv],
        "ten_nv": [u.c.ten_nv],
        "chuc_vu": [u.c.chuc_vu_raw],
        "chi_nhanh_lam": [chi_nhanh_lam_sort_expr],
        "so_cong": [u.c.so_cong],
        "la_tang_ca": [u.c.la_tang_ca],
        "dich_vu": [u.c.dich_vu],
        "so_phong": [u.c.so_phong],
        "so_luong": [u.c.so_luong],
        "type": [u.c.type],
    }

    if sort_by and sort_by in sort_map:
        sort_columns = sort_map[sort_by]
        if sort_order == 'asc':
            order_expressions.extend([col.asc().nullslast() for col in sort_columns])
        else:
            order_expressions.extend([col.desc().nullslast() for col in sort_columns])

    # Add default sort as secondary to ensure consistent ordering
    if sort_by != 'thoi_gian':
        order_expressions.extend([desc(u.c.date_col), desc(u.c.time_col)])

    # Final query with sorting and pagination
    paginated_query = final_query.order_by(*order_expressions).offset((page - 1) * per_page).limit(per_page)
    
    records = db.execute(paginated_query).all()

    # ✅ Lấy danh sách nhân viên liên quan cho bộ lọc (chỉ chạy khi page=1)
    related_employees = []
    if page == 1:
        # Lấy tất cả các mã nhân viên (ma_nv) từ các bản ghi đã lọc (không phân trang)
        employee_codes_query = select(u.c.ma_nv, u.c.ten_nv).distinct()

        # Áp dụng các bộ lọc tương tự như trên, NGOẠI TRỪ filter_nhan_vien
        # để có được danh sách đầy đủ cho dropdown
        if filter_type: employee_codes_query = employee_codes_query.where(u.c.type == filter_type)
        if filter_date:
            try:
                parsed_date = datetime.strptime(filter_date, "%Y-%m-%d").date()
                employee_codes_query = employee_codes_query.where(u.c.date_col == parsed_date)
            except ValueError: pass
        if filter_cn_lam: employee_codes_query = employee_codes_query.where(u.c.chi_nhanh_lam == filter_cn_lam)
        # ... có thể thêm các filter khác nếu cần ...

        related_employee_rows = db.execute(employee_codes_query.order_by(u.c.ten_nv)).all()
        related_employees = [{"code": row.ma_nv, "name": row.ten_nv} for row in related_employee_rows]


    # Format results
    combined_results = []
    for rec in records:
        ghi_chu_text = rec.ghi_chu or ""
        if rec.la_tang_ca and rec.type == 'Điểm danh':
            ghi_chu_text = re.sub(r'Tăng ca\s*\.?\s*', '', ghi_chu_text, flags=re.IGNORECASE).strip()
        
        dt = datetime.combine(rec.date_col, rec.time_col)

        combined_results.append({
            "id": rec.id,
            "type": rec.type,
            "thoi_gian": dt.strftime("%d/%m/%Y %H:%M:%S"),
            "nguoi_thuc_hien": rec.nguoi_thuc_hien,
            "ten_nguoi_thuc_hien": rec.ten_nguoi_thuc_hien,
            "ma_nv": rec.ma_nv,
            "ten_nv": rec.ten_nv,
            "chuc_vu": map_role_to_vietnamese(rec.chuc_vu_raw),
            "chi_nhanh_lam": rec.chi_nhanh_lam,
            "chi_nhanh_chinh": rec.chi_nhanh_chinh or rec.employee_branch,
            "so_cong": rec.so_cong,
            "tang_ca": rec.la_tang_ca,
            "ghi_chu": ghi_chu_text,
            "ghi_chu_raw": rec.ghi_chu or "",
            "dich_vu": rec.dich_vu or "",
            "so_phong": rec.so_phong or "",
            "so_luong": rec.so_luong or "",
        })

    return JSONResponse(content={
        "records": combined_results,
        "currentPage": page,
        "totalPages": total_pages,
        "totalRecords": total_records,
        "relatedEmployees": related_employees, # Trả về danh sách nhân viên cho bộ lọc
    })

def _auto_adjust_worksheet_columns(worksheet):
    """Helper function to adjust column widths of a worksheet."""
    for i, column_cells in enumerate(worksheet.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        # Also check header length
        if worksheet.cell(row=1, column=i).value:
            max_length = len(str(worksheet.cell(row=1, column=i).value))

        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

@app.get("/api/tasks/export-excel")
async def export_tasks_to_excel(
    request: Request,
    db: Session = Depends(get_db),
    chi_nhanh: str = "",
    search: str = "",
    trang_thai: str = "",
    han_hoan_thanh: str = "",
):
    """
    API để xuất danh sách công việc đã lọc ra file Excel.
    """
    user_data = request.session.get("user")
    if not user_data:
        raise HTTPException(status_code=403, detail="Bạn không có quyền truy cập.")

    tasks_query = _get_filtered_tasks_query(db, user_data, chi_nhanh, search, trang_thai, han_hoan_thanh)

    # Sắp xếp tương tự trang chủ
    order = { "Quá hạn": 0, "Đang chờ": 1, "Hoàn thành": 2, "Đã xoá": 3 }
    far_future = VN_TZ.localize(datetime(2099, 12, 31))
    rows_all = tasks_query.all()
    rows_all.sort(key=lambda t: (
        order.get(t.trang_thai, 99),
        t.han_hoan_thanh or far_future
    ))

    if not rows_all:
        return Response(status_code=204)

    data_for_export = []
    for t in rows_all:
        # Dữ liệu đã được đồng bộ ở đầu route, chỉ cần lấy trực tiếp

        data_for_export.append({
            "ID": t.id,
            "Chi Nhánh": t.chi_nhanh,
            "Phòng": t.phong,
            "Mô Tả": t.mo_ta,
            "Ngày Tạo": format_datetime_display(t.ngay_tao, with_time=True),
            "Hạn Hoàn Thành": format_datetime_display(t.han_hoan_thanh, with_time=False),
            "Trạng Thái": t.trang_thai,
            "Người Tạo": t.nguoi_tao,
            "Người Thực Hiện": t.nguoi_thuc_hien or "",
            "Ngày Hoàn Thành": format_datetime_display(t.ngay_hoan_thanh, with_time=True) if t.ngay_hoan_thanh else "",
            "Ghi Chú": t.ghi_chu or "",
        })

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CongViec"

    headers = list(data_for_export[0].keys())
    ws.append(headers)
    for row_data in data_for_export:
        ws.append(list(row_data.values()))
    _auto_adjust_worksheet_columns(ws)

    wb.save(output)
    output.seek(0)

    filename = f"danh_sach_cong_viec_{datetime.now(VN_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

@app.get("/api/attendance/export-excel")
async def export_attendance_results_to_excel(
    request: Request,
    db: Session = Depends(get_db),
    filter_type: Optional[str] = None,
    filter_date: Optional[str] = None,
    filter_nhan_vien: Optional[str] = None,
    filter_chuc_vu: Optional[str] = None,
    filter_cn_lam: Optional[str] = None,
    filter_so_cong: Optional[float] = None,
    filter_tang_ca: Optional[str] = None,
    filter_ghi_chu: Optional[str] = None,
    filter_nguoi_thuc_hien: Optional[str] = None,
    filter_dich_vu: Optional[str] = None,
    filter_so_phong: Optional[str] = None,
):
    """
    API to export filtered attendance results to an Excel file with separate sheets
    for attendance and service records.
    """
    user = request.session.get("user")
    allowed_roles = ['letan', 'quanly', 'ktv', 'admin', 'boss']
    if not user or user.get("role") not in allowed_roles:
        raise HTTPException(status_code=403, detail="Bạn không có quyền truy cập chức năng này.")
    
    checker_code = user.get("code")
    user_role = user.get("role")
    if not checker_code:
        raise HTTPException(status_code=403, detail="Không tìm thấy mã người dùng.")
    
    EmployeeUser = aliased(User, name="employee_user")
    CheckerUser = aliased(User, name="checker_user")

    # Base query for AttendanceRecord
    att_q = select(
        AttendanceRecord.id,
        AttendanceRecord.ngay_diem_danh.label("date_col"), AttendanceRecord.gio_diem_danh.label("time_col"),
        AttendanceRecord.nguoi_diem_danh.label("nguoi_thuc_hien"), CheckerUser.name.label("ten_nguoi_thuc_hien"),
        AttendanceRecord.ma_nv, AttendanceRecord.ten_nv, EmployeeUser.role.label("chuc_vu_raw"),
        AttendanceRecord.chi_nhanh_lam, AttendanceRecord.chi_nhanh_chinh, EmployeeUser.branch.label("employee_branch"),
        AttendanceRecord.so_cong_nv.label("so_cong"), AttendanceRecord.la_tang_ca, AttendanceRecord.ghi_chu
    ).join(
        EmployeeUser, EmployeeUser.code == AttendanceRecord.ma_nv, isouter=True
    ).join(
        CheckerUser, CheckerUser.code == AttendanceRecord.nguoi_diem_danh, isouter=True
    )
    
    # Base query for ServiceRecord
    svc_q = select(
        ServiceRecord.id,
        ServiceRecord.ngay_cham.label("date_col"), ServiceRecord.gio_cham.label("time_col"),
        ServiceRecord.nguoi_cham.label("nguoi_thuc_hien"), CheckerUser.name.label("ten_nguoi_thuc_hien"),
        ServiceRecord.ma_nv, ServiceRecord.ten_nv, EmployeeUser.role.label("chuc_vu_raw"),
        ServiceRecord.chi_nhanh_lam, ServiceRecord.chi_nhanh_chinh, EmployeeUser.branch.label("employee_branch"),
        ServiceRecord.la_tang_ca, ServiceRecord.ghi_chu,
        ServiceRecord.dich_vu, ServiceRecord.so_phong, ServiceRecord.so_luong
    ).join(
        EmployeeUser, EmployeeUser.code == ServiceRecord.ma_nv, isouter=True
    ).join(
        CheckerUser, CheckerUser.code == ServiceRecord.nguoi_cham, isouter=True
    )
    
    # Role-based filtering
    if user_role == "letan":
        att_q = att_q.where(or_(AttendanceRecord.nguoi_diem_danh == checker_code, AttendanceRecord.ma_nv == checker_code))
        svc_q = svc_q.where(ServiceRecord.nguoi_cham == checker_code)
    elif user_role not in ["admin", "boss"]:
        att_q = att_q.where(AttendanceRecord.nguoi_diem_danh == checker_code)
        svc_q = svc_q.where(ServiceRecord.nguoi_cham == checker_code)
    
    def _apply_common_filters(query, is_att_query=False):
        if filter_date:
            try:
                parsed_date = datetime.strptime(filter_date, "%Y-%m-%d").date()
                query = query.where(query.selected_columns.date_col == parsed_date)
            except (ValueError, TypeError): pass
        if filter_nhan_vien:
            query = query.where(or_(
                query.selected_columns.ma_nv.ilike(f"%{filter_nhan_vien}%"),
                query.selected_columns.ten_nv.ilike(f"%{filter_nhan_vien}%")
            ))
        if filter_chuc_vu:
            matching_roles = [role for role, vn_role in ROLE_MAP.items() if filter_chuc_vu.lower() in vn_role.lower()]
            if matching_roles:
                query = query.where(query.selected_columns.chuc_vu_raw.in_(matching_roles))
        if filter_cn_lam:
            query = query.where(query.selected_columns.chi_nhanh_lam.ilike(f"%{filter_cn_lam}%"))
        if is_att_query and filter_so_cong is not None:
            query = query.where(query.selected_columns.so_cong == filter_so_cong)
        if filter_tang_ca and filter_tang_ca != 'all':
            is_overtime = filter_tang_ca == 'yes'
            query = query.where(query.selected_columns.la_tang_ca == is_overtime)
        if filter_ghi_chu:
            query = query.where(query.selected_columns.ghi_chu.ilike(f"%{filter_ghi_chu}%"))
        if filter_nguoi_thuc_hien and user_role in ['admin', 'boss']:
            query = query.where(or_(
                query.selected_columns.nguoi_thuc_hien.ilike(f"%{filter_nguoi_thuc_hien}%"),
                query.selected_columns.ten_nguoi_thuc_hien.ilike(f"%{filter_nguoi_thuc_hien}%")
            ))
        if not is_att_query:
            if filter_dich_vu:
                query = query.where(query.selected_columns.dich_vu.ilike(f"%{filter_dich_vu}%"))
            if filter_so_phong:
                query = query.where(query.selected_columns.so_phong.ilike(f"%{filter_so_phong}%"))
        return query

    # Apply filters to each query
    filtered_att_q = _apply_common_filters(att_q, is_att_query=True)
    filtered_svc_q = _apply_common_filters(svc_q, is_att_query=False)

    # Fetch all records for both types, no pagination
    att_records = db.execute(filtered_att_q.order_by(desc(att_q.selected_columns.date_col), desc(att_q.selected_columns.time_col))).all()
    svc_records = db.execute(filtered_svc_q.order_by(desc(svc_q.selected_columns.date_col), desc(svc_q.selected_columns.time_col))).all()

    # Prepare data for Attendance DataFrame
    att_data_for_df = []
    for rec in att_records:
        ghi_chu_text = rec.ghi_chu or ""
        if rec.la_tang_ca:
            ghi_chu_text = re.sub(r'Tăng ca\s*\.?\s*', '', ghi_chu_text, flags=re.IGNORECASE).strip()
        dt = datetime.combine(rec.date_col, rec.time_col)
        att_data_for_df.append({
            "Thời gian": dt.strftime("%d/%m/%Y %H:%M:%S"),
            "Người thực hiện": f"{rec.ten_nguoi_thuc_hien} ({rec.nguoi_thuc_hien})" if rec.ten_nguoi_thuc_hien else rec.nguoi_thuc_hien,
            "Mã NV": rec.ma_nv, "Tên NV": rec.ten_nv, "Chức vụ": map_role_to_vietnamese(rec.chuc_vu_raw),
            "CN Làm": rec.chi_nhanh_lam, "CN Chính": rec.chi_nhanh_chinh or rec.employee_branch,
            "Tăng ca": "Có" if rec.la_tang_ca else "Không", "Số công": rec.so_cong,
            "Ghi chú": ghi_chu_text,
        })

    # Prepare data for Service DataFrame
    svc_data_for_df = []
    for rec in svc_records:
        ghi_chu_text = rec.ghi_chu or ""
        dt = datetime.combine(rec.date_col, rec.time_col)
        svc_data_for_df.append({
            "Thời gian": dt.strftime("%d/%m/%Y %H:%M:%S"),
            "Người thực hiện": f"{rec.ten_nguoi_thuc_hien} ({rec.nguoi_thuc_hien})" if rec.ten_nguoi_thuc_hien else rec.nguoi_thuc_hien,
            "Mã NV": rec.ma_nv, "Tên NV": rec.ten_nv, "Chức vụ": map_role_to_vietnamese(rec.chuc_vu_raw),
            "CN Làm": rec.chi_nhanh_lam, "CN Chính": rec.chi_nhanh_chinh or rec.employee_branch,
            "Tăng ca": "Có" if rec.la_tang_ca else "Không",
            "Dịch vụ": rec.dich_vu or "", "Số phòng": rec.so_phong or "", "Số lượng": rec.so_luong or "",
            "Ghi chú": ghi_chu_text,
        })

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet

    # --- Sheet 1: Điểm danh ---
    if att_data_for_df:
        ws_att = wb.create_sheet(title="Điểm danh")
        headers_att = list(att_data_for_df[0].keys())
        ws_att.append(headers_att)
        for row_data in att_data_for_df:
            ws_att.append(list(row_data.values()))
        _auto_adjust_worksheet_columns(ws_att)

    # --- Sheet 2: Chấm Dịch Vụ ---
    if svc_data_for_df:
        ws_svc = wb.create_sheet(title="Chấm Dịch Vụ")
        headers_svc = list(svc_data_for_df[0].keys())
        ws_svc.append(headers_svc)
        for row_data in svc_data_for_df:
            ws_svc.append(list(row_data.values()))
        _auto_adjust_worksheet_columns(ws_svc)

    if not wb.sheetnames: # If no data was added
        return Response(status_code=204)

    wb.save(output)
    output.seek(0)

    filename = f"ket_qua_diem_danh_{datetime.now(VN_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

class RecordToDelete(BaseModel):
    id: int
    type: str

class BatchDeletePayload(BaseModel):
    records: list[RecordToDelete]

@app.post("/api/attendance/records/batch-delete", response_class=JSONResponse)
async def batch_delete_records(
    request: Request,
    payload: BatchDeletePayload,
    db: Session = Depends(get_db)
):
    user = request.session.get("user")
    if not user or user.get("role") not in ["admin", "boss"]:
        raise HTTPException(status_code=403, detail="Chỉ admin hoặc boss mới có quyền xóa hàng loạt.")

    deleted_count = 0
    try:
        for record_info in payload.records:
            if record_info.type == 'attendance':
                record = db.query(AttendanceRecord).filter(AttendanceRecord.id == record_info.id).first()
            elif record_info.type == 'service':
                record = db.query(ServiceRecord).filter(ServiceRecord.id == record_info.id).first()
            else:
                continue

            if record:
                db.delete(record)
                deleted_count += 1
        db.commit()
        return JSONResponse({"status": "success", "message": f"Đã xóa {deleted_count} bản ghi.", "deleted_count": deleted_count})
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Lỗi cơ sở dữ liệu: {e}")

# --- Manual Absence Check for Admin/Boss ---
class AbsenceCheckPayload(BaseModel):
    check_date: str

@app.post("/api/attendance/run-absence-check", response_class=JSONResponse)
async def trigger_absence_check(
    request: Request,
    payload: AbsenceCheckPayload,
    background_tasks: BackgroundTasks,
):
    user = request.session.get("user")
    if not user or user.get("role") not in ["admin", "boss"]:
        raise HTTPException(status_code=403, detail="Bạn không có quyền thực hiện chức năng này.")

    try:
        target_date = datetime.strptime(payload.check_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Định dạng ngày không hợp lệ. Vui lòng dùng YYYY-MM-DD.")

    # Chạy tác vụ trong nền để không block request
    background_tasks.add_task(run_daily_absence_check, target_date=target_date)

    return JSONResponse(
        content={
            "status": "success",
            "message": f"Đã yêu cầu cập nhật điểm danh vắng cho ngày {target_date.strftime('%d/%m/%Y')}. Quá trình sẽ chạy trong nền."
        }
    )


# --- Manual Record Management for Admin/Boss ---

def parse_form_datetime(dt_str: str) -> Optional[datetime]:
    """Hàm helper để parse datetime từ form của modal sửa/thêm."""
    if not dt_str:
        return None
    try:
        # Frontend sends 'dd/mm/yyyy HH:MM' after conversion
        return VN_TZ.localize(datetime.strptime(dt_str, "%d/%m/%Y %H:%M"))
    except (ValueError, TypeError):
        return None

@app.get("/api/users/search-checkers", response_class=JSONResponse)
def search_checkers(q: str = "", db: Session = Depends(get_db)):
    """API để tìm kiếm người dùng có quyền điểm danh (lễ tân, ql, ktv, admin, boss)."""
    if not q:
        return JSONResponse(content=[])
    
    search_pattern = f"%{q}%"
    allowed_roles = ["letan", "quanly", "ktv", "admin", "boss"]
    
    users = db.query(User).filter(
        User.role.in_(allowed_roles),
        or_(
            User.code.ilike(search_pattern),
            User.name.ilike(search_pattern)
        )
    ).limit(20).all()
    
    user_list = [
        {"code": user.code, "name": user.name}
        for user in users
    ]
    return JSONResponse(content=user_list)

@app.post("/api/attendance/manual-record", response_class=JSONResponse)
async def create_manual_record(
    request: Request,
    db: Session = Depends(get_db),
    record_type: str = Form(...),
    ma_nv: str = Form(...),
    thoi_gian: str = Form(...),
    nguoi_thuc_hien: Optional[str] = Form(None),
    chi_nhanh_lam: Optional[str] = Form(None),
    la_tang_ca: bool = Form(False),
    ghi_chu: Optional[str] = Form(""),
    so_cong_nv: Optional[float] = Form(1.0),
    dich_vu: Optional[str] = Form(""),
    so_phong: Optional[str] = Form(""),
    so_luong: Optional[str] = Form(""),
):
    user = request.session.get("user")
    if not user or user.get("role") not in ["admin", "boss"]:
        raise HTTPException(status_code=403, detail="Chỉ admin hoặc boss mới có quyền thực hiện.")

    employee = db.query(User).filter(User.code == ma_nv).first()
    if not employee:
        raise HTTPException(status_code=404, detail=f"Không tìm thấy nhân viên với mã: {ma_nv}")

    dt_obj = parse_form_datetime(thoi_gian)
    if not dt_obj:
        raise HTTPException(status_code=400, detail="Định dạng thời gian không hợp lệ. Cần: dd/mm/yyyy HH:MM")

    # ✅ Logic xác định tăng ca cho quản lý và KTV
    final_la_tang_ca = la_tang_ca
    if employee.role in ["ktv", "quanly"]:
        final_la_tang_ca = (so_cong_nv or 0) > 1.0

    is_ktv_or_quanly = employee.role in ["ktv", "quanly"]
    final_chi_nhanh_lam = chi_nhanh_lam if not is_ktv_or_quanly else employee.branch
    if not final_chi_nhanh_lam:
        raise HTTPException(status_code=400, detail="Không thể xác định chi nhánh làm việc.")

    final_nguoi_thuc_hien = nguoi_thuc_hien if not is_ktv_or_quanly else employee.code
    if not final_nguoi_thuc_hien:
        # Nếu không có người thực hiện, mặc định là người đang đăng nhập
        final_nguoi_thuc_hien = user.get("code")

    try:
        if record_type == 'attendance':
            new_record = AttendanceRecord(
                ngay_diem_danh=dt_obj.date(),
                gio_diem_danh=dt_obj.time(),
                nguoi_diem_danh=final_nguoi_thuc_hien,
                ma_nv=ma_nv,
                ten_nv=employee.name,
                chi_nhanh_chinh=employee.branch,
                chi_nhanh_lam=final_chi_nhanh_lam,
                la_tang_ca=final_la_tang_ca,
                so_cong_nv=so_cong_nv or 1.0,
                ghi_chu=ghi_chu
            )
            db.add(new_record)
        elif record_type == 'service':
            new_record = ServiceRecord(
                ngay_cham=dt_obj.date(),
                gio_cham=dt_obj.time(),
                nguoi_cham=final_nguoi_thuc_hien,
                ma_nv=ma_nv,
                ten_nv=employee.name,
                chi_nhanh_chinh=employee.branch,
                chi_nhanh_lam=final_chi_nhanh_lam,
                la_tang_ca=final_la_tang_ca,
                dich_vu=dich_vu,
                so_phong=so_phong,
                so_luong=so_luong,
                ghi_chu=ghi_chu
            )
            db.add(new_record)
        else:
            raise HTTPException(status_code=400, detail="Loại bản ghi không hợp lệ.")
        
        db.commit()
        return JSONResponse({"status": "success", "message": "Đã thêm bản ghi thành công."})
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Lỗi cơ sở dữ liệu: {e}")

@app.post("/api/attendance/manual-record/{record_type}/{record_id}", response_class=JSONResponse)
async def update_manual_record(
    record_type: str,
    record_id: int,
    request: Request,
    db: Session = Depends(get_db),
    ma_nv: str = Form(...),
    thoi_gian: str = Form(...),
    nguoi_thuc_hien: Optional[str] = Form(None),
    chi_nhanh_lam: Optional[str] = Form(None),
    la_tang_ca: bool = Form(False),
    ghi_chu: Optional[str] = Form(""),
    so_cong_nv: Optional[float] = Form(1.0),
    dich_vu: Optional[str] = Form(""),
    so_phong: Optional[str] = Form(""),
    so_luong: Optional[str] = Form(""),
):
    user = request.session.get("user")
    if not user or user.get("role") not in ["admin", "boss"]:
        raise HTTPException(status_code=403, detail="Chỉ admin hoặc boss mới có quyền thực hiện.")

    employee = db.query(User).filter(User.code == ma_nv).first()
    if not employee:
        raise HTTPException(status_code=404, detail=f"Không tìm thấy nhân viên với mã: {ma_nv}")

    dt_obj = parse_form_datetime(thoi_gian)
    if not dt_obj:
        raise HTTPException(status_code=400, detail="Định dạng thời gian không hợp lệ. Cần: dd/mm/yyyy HH:MM")

    # ✅ Logic xác định tăng ca cho quản lý và KTV
    final_la_tang_ca = la_tang_ca
    if employee.role in ["ktv", "quanly"]:
        final_la_tang_ca = (so_cong_nv or 0) > 1.0

    is_ktv_or_quanly = employee.role in ["ktv", "quanly"]
    final_chi_nhanh_lam = chi_nhanh_lam if not is_ktv_or_quanly else employee.branch
    if not final_chi_nhanh_lam:
        raise HTTPException(status_code=400, detail="Không thể xác định chi nhánh làm việc.")

    final_nguoi_thuc_hien = nguoi_thuc_hien if not is_ktv_or_quanly else employee.code
    if not final_nguoi_thuc_hien:
        # Khi cập nhật, người thực hiện phải luôn được cung cấp cho các role khác KTV/QL
        raise HTTPException(status_code=400, detail="Không thể xác định người thực hiện.")

    try:
        if record_type == 'attendance':
            record = db.query(AttendanceRecord).filter(AttendanceRecord.id == record_id).first()
            if not record:
                raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi điểm danh.")
            
            record.ngay_diem_danh = dt_obj.date()
            record.gio_diem_danh = dt_obj.time()
            record.nguoi_diem_danh = final_nguoi_thuc_hien
            record.ma_nv = ma_nv
            record.ten_nv = employee.name
            record.chi_nhanh_chinh = employee.branch
            record.chi_nhanh_lam = final_chi_nhanh_lam
            record.la_tang_ca = final_la_tang_ca
            record.so_cong_nv = so_cong_nv or 1.0
            record.ghi_chu = ghi_chu

        elif record_type == 'service':
            record = db.query(ServiceRecord).filter(ServiceRecord.id == record_id).first()
            if not record:
                raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi dịch vụ.")

            record.ngay_cham = dt_obj.date()
            record.gio_cham = dt_obj.time()
            record.nguoi_cham = final_nguoi_thuc_hien
            record.ma_nv = ma_nv
            record.ten_nv = employee.name
            record.chi_nhanh_chinh = employee.branch
            record.chi_nhanh_lam = final_chi_nhanh_lam
            record.la_tang_ca = final_la_tang_ca
            record.dich_vu = dich_vu
            record.so_phong = so_phong
            record.so_luong = so_luong
            record.ghi_chu = ghi_chu
        else:
            raise HTTPException(status_code=400, detail="Loại bản ghi không hợp lệ.")
        
        db.commit()
        return JSONResponse({"status": "success", "message": "Đã cập nhật bản ghi thành công."})
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Lỗi cơ sở dữ liệu: {e}")

@app.delete("/api/attendance/record/{record_type}/{record_id}", response_class=JSONResponse)
async def delete_manual_record(
    record_type: str,
    record_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user = request.session.get("user")
    if not user or user.get("role") not in ["admin", "boss"]:
        raise HTTPException(status_code=403, detail="Chỉ admin hoặc boss mới có quyền thực hiện.")
        
    try:
        if record_type == 'attendance':
            record = db.query(AttendanceRecord).filter(AttendanceRecord.id == record_id).first()
        elif record_type == 'service':
            record = db.query(ServiceRecord).filter(ServiceRecord.id == record_id).first()
        else:
            raise HTTPException(status_code=400, detail="Loại bản ghi không hợp lệ.")

        if not record:
            raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi.")
        
        db.delete(record)
        db.commit()
        return JSONResponse({"status": "success", "message": "Đã xóa bản ghi thành công."})
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Lỗi cơ sở dữ liệu: {e}")

# --- QR Checkin APIs ---

@app.get("/attendance/checkin")
def attendance_checkin(request: Request, token: str, db: Session = Depends(get_db)):
    log = db.query(AttendanceLog).filter_by(token=token).first()
    if not log:
        return HTMLResponse("Token không hợp lệ!", status_code=400)

    if log.checked_in:
        return templates.TemplateResponse(
            "qr_invalid.html",
            {"request": request, "message": "Mã QR này đã được sử dụng để điểm danh và không còn hợp lệ."},
            status_code=403
        )

    user = db.query(User).filter_by(code=log.user_code).first()
    if not user:
        return HTMLResponse("Không tìm thấy user!", status_code=400)

    user_data = {
        "code": user.code,
        "role": user.role,
        "branch": user.branch,
        "name": user.name
    }

    request.session.pop("user", None)
    request.session["pending_user"] = user_data
    role = user_data.get("role", "")

    return templates.TemplateResponse("attendance.html", {
        "request": request,
        "role": role,
        "branch_id": user.branch,
        "csrf_token": get_csrf_token(request),
        "user": user_data,
        "login_code": user.code,
        "token": token
    })

@app.post("/attendance/checkin_success")
async def checkin_success(request: Request, db: Session = Depends(get_db)):
    # ✅ Boss & Admin: bỏ qua điểm danh, vào thẳng hệ thống
    pending = request.session.get("pending_user") or request.session.get("user")
    if pending and pending.get("role") in ["boss", "admin"]:
        request.session["user"] = dict(pending)  # copy toàn bộ thông tin
        request.session["after_checkin"] = "choose_function"
        request.session.pop("pending_user", None)
        return JSONResponse({"success": True, "redirect_to": "/choose-function"})

    # ✅ Các role khác: tiếp tục xử lý điểm danh
    data = await request.json()
    token = data.get("token")
    user_code = None
    work_date, shift = get_current_work_shift()
    log = None
    user = None

    # Luồng 1: Điểm danh qua QR code (desktop có token)
    if token:
        log = db.query(AttendanceLog).filter_by(token=token).first()
        if not log:
            return JSONResponse({"success": False, "error": "Token không hợp lệ"}, status_code=400)

        user_code = log.user_code
        user = db.query(User).filter_by(code=user_code).first()
        # ✅ Xác định shift_value nhất quán theo role
        shift_value = _get_log_shift_for_user(user.role, log.shift) if user else log.shift

    # Luồng 2: Mobile (không có token)
    else:
        pending = request.session.get("pending_user")
        if not pending or not pending.get("code"):
            return JSONResponse(
                {"success": False, "error": "Không tìm thấy pending_user trong session."},
                status_code=403
            )

        user_code = pending["code"]
        user = db.query(User).filter_by(code=user_code).first()
        if not user:
            return JSONResponse({"success": False, "error": "Không tìm thấy người dùng"}, status_code=404)

        # ✅ shift_value: ktv/quanly = None, còn lại theo ca
        shift_value = _get_log_shift_for_user(user.role, shift)

        # Query log theo shift_value (tránh sinh 2 log)
        log = db.query(AttendanceLog).filter_by(
            user_code=user_code,
            date=work_date,
            shift=shift_value
        ).first()

        # Nếu chưa có log thì tạo mới
        if not log:
            log = AttendanceLog(
                user_code=user_code,
                date=work_date,
                shift=shift_value,
                checked_in=False,
                token=secrets.token_urlsafe(24)
            )
            db.add(log)
            db.commit()

    # ✅ Cập nhật check-in
    if not log:
        return JSONResponse({"success": False, "error": "Không tìm thấy bản ghi điểm danh."}, status_code=404)

    log.checked_in = True
    db.commit()

    # ✅ Cập nhật session chính thức
    if user:
        request.session["user"] = {
            "code": user.code,
            "role": user.role,
            "branch": user.branch,
            "name": user.name
        }
        request.session["after_checkin"] = "choose_function"
        request.session.pop("pending_user", None)
        return JSONResponse({"success": True, "redirect_to": "/choose-function"})

    return JSONResponse({"success": False, "error": "Không tìm thấy người dùng"}, status_code=404)

@app.get("/attendance/checkin_status")
async def checkin_status(request: Request, token: str, db: Session = Depends(get_db)):
    log = db.query(AttendanceLog).filter_by(token=token).first()
    if not log:
        return JSONResponse(content={"checked_in": False})

    if log.checked_in:
        user = db.query(User).filter_by(code=log.user_code).first()
        if user:
            # Đăng nhập cho user ở session của máy tính
            request.session["user"] = {
                "code": user.code,
                "role": user.role,
                "branch": user.branch,
                "name": user.name,
            }
            request.session.pop("pending_user", None)
            return JSONResponse(content={"checked_in": True, "redirect_to": "choose-function"})

    return JSONResponse(content={"checked_in": False})

@app.get("/favicon.ico")
def favicon():
    # Nếu có file favicon.ico trong static thì trả về file đó
    favicon_path = os.path.join("static", "favicon.ico")
    if os.path.exists(favicon_path):
        return FileResponse(favicon_path)
    # Nếu không có, trả về 1x1 PNG trắng
    png_data = b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\nIDATx\xdac\xf8\xff\xff?\x00\x05\xfe\x02\xfeA\x0b\x0c\x00\x00\x00\x00IEND\xaeB`\x82'
    return Response(content=png_data, media_type="image/png")
